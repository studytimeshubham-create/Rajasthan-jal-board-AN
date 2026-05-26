import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
import os
from datetime import datetime

def get_frame(parent, fc, utils, be, admin) -> ttk.Frame:
    frame = ttk.Frame(parent)
    
    # Notebook for tabs
    notebook = ttk.Notebook(frame)
    notebook.pack(fill="both", expand=True)
    
    # Tab 1: Search / View Consumers
    search_tab = ttk.Frame(notebook, padding=10)
    notebook.add(search_tab, text="🔍 Search / View")
    build_search_tab(search_tab, fc, utils, be, admin)
    
    # Tab 2: Add Consumer
    add_tab = ttk.Frame(notebook, padding=10)
    notebook.add(add_tab, text="➕ Add Consumer")
    build_add_tab(add_tab, fc, utils, be, admin, notebook)
    
    # Tab 3: Bulk Import CSD
    import_tab = ttk.Frame(notebook, padding=10)
    notebook.add(import_tab, text="📥 Bulk Import")
    build_import_tab(import_tab, fc, utils, be, admin)
    
    # Tab 4: Export CSD
    export_tab = ttk.Frame(notebook, padding=10)
    notebook.add(export_tab, text="📤 Export CSD")
    build_export_tab(export_tab, fc, utils, be, admin)
    
    # Tab 5: Meter Replacement
    replace_tab = ttk.Frame(notebook, padding=10)
    notebook.add(replace_tab, text="🔧 Meter Replacement")
    build_replacement_tab(replace_tab, fc, utils, be, admin)
    
    return frame

# ----------------------------------------------------
# TAB 1: Search & View
# ----------------------------------------------------
def build_search_tab(tab, fc, utils, be, admin):
    # Top Search Bar
    s_frame = ttk.Frame(tab, padding=5)
    s_frame.pack(fill="x", pady=(0, 10))
    
    ttk.Label(s_frame, text="Search Query (CIN or Meter Serial):").pack(side="left", padx=5)
    search_ent = ttk.Entry(s_frame, width=30)
    search_ent.pack(side="left", padx=5)
    
    # Filter dropdowns
    ttk.Label(s_frame, text="Zone:").pack(side="left", padx=5)
    zone_var = tk.StringVar(value="All")
    zone_cb = ttk.Combobox(s_frame, textvariable=zone_var, values=["All"] + [str(z) for z in utils.ZONE_RANGE], width=6, state="readonly")
    zone_cb.pack(side="left", padx=5)
    
    ttk.Label(s_frame, text="Status:").pack(side="left", padx=5)
    status_var = tk.StringVar(value="All")
    status_cb = ttk.Combobox(s_frame, textvariable=status_var, values=["All"] + utils.CONSUMER_STATUS_OPTIONS, width=10, state="readonly")
    status_cb.pack(side="left", padx=5)
    
    tree_columns = ("cin_no", "name", "zone", "category", "meter_serial_no", "last_reading", "outstanding_balance", "consumer_status")
    
    def perform_search():
        query = search_ent.get().strip()
        zone_f = zone_cb.get()
        status_f = status_cb.get()
        
        # Clear tree
        for row in tree.get_children():
            tree.delete(row)
            
        def fetch():
            filters = {"is_active": True}
            if zone_f != "All":
                filters["zone"] = int(zone_f)
            if status_f != "All":
                filters["status"] = status_f
                
            consumers = fc.list_consumers(filters)
            
            # Local client-side text filtering if query set
            if query:
                q = query.lower()
                consumers = [c for c in consumers if q in c["cin_no"].lower() or q in str(c.get("meter_serial_no", "")).lower() or q in c["name"].lower()]
            return consumers
            
        def on_fetched(consumers):
            for c in consumers:
                tree.insert("", "end", values=(
                    c["cin_no"],
                    c["name"],
                    c["zone"],
                    c["category"],
                    c.get("meter_serial_no", ""),
                    f"{c.get('last_reading', 0.0):.2f} KL",
                    utils.format_currency(c.get("outstanding_balance", 0.0)),
                    c.get("consumer_status", "Active")
                ))
                
        utils.run_in_thread(fetch, callback=on_fetched, widget=tab)

    search_btn = ttk.Button(s_frame, text="Search / Refresh", command=perform_search)
    search_btn.pack(side="left", padx=10)
    
    # Treeview container
    t_frame = ttk.Frame(tab)
    t_frame.pack(fill="both", expand=True)
    
    tree = ttk.Treeview(t_frame, columns=tree_columns, show="headings")
    tree.heading("cin_no", text="CIN Number")
    tree.heading("name", text="Consumer Name")
    tree.heading("zone", text="Zone")
    tree.heading("category", text="Category")
    tree.heading("meter_serial_no", text="Meter Serial")
    tree.heading("last_reading", text="Last Reading")
    tree.heading("outstanding_balance", text="Outstanding")
    tree.heading("consumer_status", text="Status")
    
    tree.column("cin_no", width=110, anchor="center")
    tree.column("name", width=180, anchor="w")
    tree.column("zone", width=50, anchor="center")
    tree.column("category", width=90, anchor="center")
    tree.column("meter_serial_no", width=100, anchor="center")
    tree.column("last_reading", width=100, anchor="e")
    tree.column("outstanding_balance", width=110, anchor="e")
    tree.column("consumer_status", width=80, anchor="center")
    
    scrollbar = ttk.Scrollbar(t_frame, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    
    tree.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    # Double-click logic
    def on_row_double_click(event):
        item = tree.selection()
        if not item:
            return
        cin_no = tree.item(item[0], "values")[0]
        show_consumer_detail_popup(cin_no, fc, utils, be, admin, tab, perform_search)
        
    tree.bind("<Double-1>", on_row_double_click)
    
    # Initial load
    perform_search()

def show_consumer_detail_popup(cin_no, fc, utils, be, admin, parent_widget, refresh_callback):
    popup = tk.Toplevel(parent_widget)
    popup.title(f"Consumer Detail — {cin_no}")
    popup.geometry("750x550")
    popup.grab_set()
    
    # Inner Tabs: Profile Details and History Ledger
    inner_notebook = ttk.Notebook(popup)
    inner_notebook.pack(fill="both", expand=True, padx=10, pady=10)
    
    profile_tab = ttk.Frame(inner_notebook, padding=10)
    inner_notebook.add(profile_tab, text="👤 Profile Info")
    
    ledger_tab = ttk.Frame(inner_notebook, padding=10)
    inner_notebook.add(ledger_tab, text="📖 History Ledger")
    
    consumer_data = [None]  # Local cache wrapper
    
    def load_data():
        def fetch():
            c = fc.get_consumer(cin_no)
            r = fc.get_readings_for_cycle("", cin_no) # fetch readings history (non-cycle specific)
            p = fc.get_payments_for_consumer(cin_no)
            a = fc.get_adjustments_for_consumer(cin_no)
            m = fc.get_meter_replacement_history(cin_no)
            return c, r, p, a, m
            
        def on_fetched(res):
            c, readings, payments, adjustments, replacements = res
            consumer_data[0] = c
            if not c:
                messagebox.showerror("Error", "Failed to retrieve consumer data.", parent=popup)
                popup.destroy()
                return
                
            # Build Profile View
            build_profile_view(profile_tab, c, fc, utils, be, admin, popup, refresh_callback, load_data)
            # Build Ledger View
            build_ledger_history_view(ledger_tab, readings, payments, adjustments, replacements, utils)
            
        utils.run_in_thread(fetch, callback=on_fetched, widget=popup)
        
    load_data()

def build_profile_view(tab, c, fc, utils, be, admin, popup, refresh_callback, reload_callback):
    for w in tab.winfo_children():
        w.destroy()
        
    f = ttk.Frame(tab)
    f.pack(fill="both", expand=True)
    
    # 2-column key-value grid for details
    grid = ttk.Frame(f)
    grid.pack(fill="x", pady=10)
    
    details = [
        ("CIN Number:", c["cin_no"], 0, 0),
        ("Consumer Name:", c["name"], 0, 1),
        ("Zone Number:", c["zone"], 1, 0),
        ("Billing Category:", c["category"], 1, 1),
        ("Meter Size:", c["meter_size"], 2, 0),
        ("Meter Serial No:", c.get("meter_serial_no", ""), 2, 1),
        ("Contact Number:", c.get("contact_number") or "", 3, 0),
        ("Aadhaar/PHED No:", c.get("aadhaar_phed_no") or "", 3, 1),
        ("Status:", c.get("consumer_status", "Active"), 4, 0),
        ("Is Active:", "Yes" if c.get("is_active", True) else "No", 4, 1),
        ("Outstanding Balance:", utils.format_currency(c.get("outstanding_balance", 0.0)), 5, 0),
        ("Credit Balance:", utils.format_currency(c.get("credit_balance", 0.0)), 5, 1),
        ("Initial Meter Reading:", f"{c.get('initial_meter_reading', 0.0):.2f} KL", 6, 0),
        ("Last Verified Reading:", f"{c.get('last_reading', 0.0):.2f} KL", 6, 1),
        ("GPS Coordinates:", f"{c.get('address_latitude','')}, {c.get('address_longitude','')}", 7, 0),
        ("PIN Code:", c.get("address_pin_code") or "", 7, 1),
        ("Area/Location:", c.get("address_area_location") or "", 8, 0),
        ("Landmark:", c.get("address_landmark") or "", 8, 1)
    ]
    
    for lbl, val, row, col in details:
        lbl_w = ttk.Label(grid, text=lbl, font=("Segoe UI", 9, "bold"))
        lbl_w.grid(row=row, column=col*2, sticky="w", padx=10, pady=3)
        val_w = ttk.Label(grid, text=str(val))
        val_w.grid(row=row, column=col*2+1, sticky="w", padx=10, pady=3)
        
    # Custom Attributes list inside Profile
    attrs = c.get("custom_attributes", {})
    if attrs:
        attr_title = ttk.Label(f, text="Custom Attributes", font=("Segoe UI", 10, "bold"), foreground="#1a3a6b")
        attr_title.pack(anchor="w", pady=(10, 5))
        attr_grid = ttk.Frame(f)
        attr_grid.pack(fill="x")
        r_idx = 0
        for k, v in attrs.items():
            ttk.Label(attr_grid, text=f"{k}:", font=("Segoe UI", 9, "italic")).grid(row=r_idx, column=0, sticky="w", padx=10, pady=2)
            ttk.Label(attr_grid, text=str(v)).grid(row=r_idx, column=1, sticky="w", padx=10, pady=2)
            r_idx += 1
            
    # Action buttons at bottom
    btn_frame = ttk.Frame(f, padding=10)
    btn_frame.pack(side="bottom", fill="x")
    
    def edit_consumer_action():
        show_edit_consumer_dialog(c, fc, utils, admin, popup, lambda: [reload_callback(), refresh_callback()])
        
    def toggle_status_action():
        active = c.get("is_active", True)
        act_str = "Deactivate" if active else "Reactivate"
        confirm = messagebox.askyesno("Confirm Toggle", f"Are you sure you want to {act_str.lower()} consumer {c['cin_no']}?", parent=popup)
        if not confirm:
            return
            
        def update():
            if active:
                fc.deactivate_consumer(c["cin_no"], admin["name"])
            else:
                fc.reactivate_consumer(c["cin_no"], admin["name"])
                
        def done(res):
            messagebox.showinfo("Success", f"Consumer status updated successfully.", parent=popup)
            reload_callback()
            refresh_callback()
            
        utils.run_in_thread(update, callback=done, widget=popup)
        
    def print_csd():
        def run():
            html_tmpl = utils.load_pdf_template("csd_sheet")
            # Build custom attributes rows for the HTML template
            c_rows = ""
            for k, v in c.get("custom_attributes", {}).items():
                c_rows += f"<tr><td class='label'>{k}</td><td class='value' colspan='3'>{v}</td></tr>"
                
            html = html_tmpl.replace("{{cin_no}}", c["cin_no"])\
                            .replace("{{name}}", c["name"])\
                            .replace("{{zone}}", str(c["zone"]))\
                            .replace("{{category}}", c["category"])\
                            .replace("{{meter_size}}", c["meter_size"])\
                            .replace("{{meter_serial_no}}", c.get("meter_serial_no") or "")\
                            .replace("{{contact_number}}", str(c.get("contact_number") or ""))\
                            .replace("{{aadhaar_phed_no}}", c.get("aadhaar_phed_no") or "")\
                            .replace("{{consumer_status}}", c.get("consumer_status", "Active"))\
                            .replace("{{last_reading}}", f"{c.get('last_reading', 0.0):.2f}")\
                            .replace("{{apl_bpl}}", c.get("apl_bpl", "APL"))\
                            .replace("{{address}}", c.get("address_area_location", ""))\
                            .replace("{{address_landmark}}", c.get("address_landmark", ""))\
                            .replace("{{address_pin_code}}", str(c.get("address_pin_code") or ""))\
                            .replace("{{address_latitude}}", str(c.get("address_latitude") or "0.00"))\
                            .replace("{{address_longitude}}", str(c.get("address_longitude") or "0.00"))\
                            .replace("{{print_date}}", datetime.now().strftime("%d-%m-%Y"))\
                            .replace("{{cycle_period}}", "N/A")\
                            .replace("{{custom_attributes_rows}}", c_rows)
                            
            pdf_bytes = utils.render_pdf_to_bytes(html)
            temp_path = os.path.join(os.environ.get("TEMP", "."), f"csd_{c['cin_no']}.pdf")
            with open(temp_path, "wb") as f_out:
                f_out.write(pdf_bytes)
            utils.open_pdf(temp_path)
            
        utils.run_in_thread(run, widget=popup)

    ttk.Button(btn_frame, text="✏️ Edit Info", command=edit_consumer_action).pack(side="left", padx=5)
    
    toggle_text = "🔒 Deactivate" if c.get("is_active", True) else "🔓 Reactivate"
    ttk.Button(btn_frame, text=toggle_text, command=toggle_status_action).pack(side="left", padx=5)
    
    ttk.Button(btn_frame, text="🖨️ CSD Sheet PDF", command=print_csd).pack(side="right", padx=5)

def build_ledger_history_view(tab, readings, payments, adjustments, replacements, utils):
    for w in tab.winfo_children():
        w.destroy()
        
    nb = ttk.Notebook(tab)
    nb.pack(fill="both", expand=True)
    
    # 1. Readings Tab
    r_tab = ttk.Frame(nb)
    nb.add(r_tab, text="Readings")
    r_tree = ttk.Treeview(r_tab, columns=("date", "cycle", "prev", "curr", "consumption", "status"), show="headings")
    r_tree.heading("date", text="Date")
    r_tree.heading("cycle", text="Cycle ID")
    r_tree.heading("prev", text="Prev (KL)")
    r_tree.heading("curr", text="Curr (KL)")
    r_tree.heading("consumption", text="Usage (KL)")
    r_tree.heading("status", text="Status")
    r_tree.column("date", width=90)
    r_tree.column("cycle", width=100)
    r_tree.column("prev", width=80, anchor="e")
    r_tree.column("curr", width=80, anchor="e")
    r_tree.column("consumption", width=80, anchor="e")
    r_tree.column("status", width=80, anchor="center")
    
    r_scroll = ttk.Scrollbar(r_tab, orient="vertical", command=r_tree.yview)
    r_tree.configure(yscrollcommand=r_scroll.set)
    r_tree.pack(side="left", fill="both", expand=True)
    r_scroll.pack(side="right", fill="y")
    
    for r in readings:
        r_tree.insert("", "end", values=(
            r.get("reading_date", ""),
            r.get("cycle_id", ""),
            f"{r.get('previous_reading', 0.0):.2f}",
            f"{r.get('current_reading', 0.0):.2f}" if r.get("current_reading") is not None else "",
            f"{r.get('consumption', 0.0):.2f}" if r.get("consumption") is not None else "",
            r.get("status", "")
        ))
        
    # 2. Payments Tab
    p_tab = ttk.Frame(nb)
    nb.add(p_tab, text="Payments")
    p_tree = ttk.Treeview(p_tab, columns=("date", "receipt", "mode", "amount"), show="headings")
    p_tree.heading("date", text="Date")
    p_tree.heading("receipt", text="Receipt No")
    p_tree.heading("mode", text="Mode")
    p_tree.heading("amount", text="Amount")
    p_tree.column("date", width=100)
    p_tree.column("receipt", width=150)
    p_tree.column("mode", width=100, anchor="center")
    p_tree.column("amount", width=120, anchor="e")
    
    p_scroll = ttk.Scrollbar(p_tab, orient="vertical", command=p_tree.yview)
    p_tree.configure(yscrollcommand=p_scroll.set)
    p_tree.pack(side="left", fill="both", expand=True)
    p_scroll.pack(side="right", fill="y")
    
    for pay in payments:
        p_tree.insert("", "end", values=(
            pay.get("payment_date"),
            pay.get("receipt_number"),
            pay.get("payment_mode"),
            utils.format_currency(pay.get("amount", 0.0))
        ))
        
    # 3. Adjustments Tab
    a_tab = ttk.Frame(nb)
    nb.add(a_tab, text="Adjustments")
    a_tree = ttk.Treeview(a_tab, columns=("date", "type", "reason", "amount"), show="headings")
    a_tree.heading("date", text="Applied At")
    a_tree.heading("type", text="Type")
    a_tree.heading("reason", text="Reason/Note")
    a_tree.heading("amount", text="Amount")
    a_tree.column("date", width=120)
    a_tree.column("type", width=100, anchor="center")
    a_tree.column("reason", width=250)
    a_tree.column("amount", width=100, anchor="e")
    
    a_scroll = ttk.Scrollbar(a_tab, orient="vertical", command=a_tree.yview)
    a_tree.configure(yscrollcommand=a_scroll.set)
    a_tree.pack(side="left", fill="both", expand=True)
    a_scroll.pack(side="right", fill="y")
    
    for adj in adjustments:
        a_tree.insert("", "end", values=(
            utils.format_date(adj.get("applied_at")),
            adj.get("type"),
            adj.get("reason_note"),
            utils.format_currency(adj.get("amount", 0.0))
        ))
        
    # 4. Replacement Log
    m_tab = ttk.Frame(nb)
    nb.add(m_tab, text="Replacements")
    m_tree = ttk.Treeview(m_tab, columns=("date", "old_serial", "new_serial", "new_reading"), show="headings")
    m_tree.heading("date", text="Date")
    m_tree.heading("old_serial", text="Old Serial")
    m_tree.heading("new_serial", text="New Serial")
    m_tree.heading("new_reading", text="Initial Reading")
    m_tree.column("date", width=100)
    m_tree.column("old_serial", width=120)
    m_tree.column("new_serial", width=120)
    m_tree.column("new_reading", width=120, anchor="e")
    
    m_scroll = ttk.Scrollbar(m_tab, orient="vertical", command=m_tree.yview)
    m_tree.configure(yscrollcommand=m_scroll.set)
    m_tree.pack(side="left", fill="both", expand=True)
    m_scroll.pack(side="right", fill="y")
    
    for rep in replacements:
        m_tree.insert("", "end", values=(
            rep.get("replacement_date"),
            rep.get("old_meter_serial"),
            rep.get("new_meter_serial"),
            f"{rep.get('new_initial_reading', 0.0):.2f} KL"
        ))

# ----------------------------------------------------
# TAB 2: Add Consumer
# ----------------------------------------------------
def build_add_tab(tab, fc, utils, be, admin, notebook):
    # Form Frame wrapped in Canvas + Scrollbar for long contents
    canvas = tk.Canvas(tab, borderwidth=0, highlightthickness=0)
    scrollbar = ttk.Scrollbar(tab, orient="vertical", command=canvas.yview)
    scroll_frame = ttk.Frame(canvas)
    
    scroll_frame.bind(
        "<Configure>",
        lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
    )
    
    canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
    canvas.configure(yscrollcommand=scrollbar.set)
    
    canvas.pack(side="left", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    # Form layout
    f = ttk.Frame(scroll_frame, padding=10)
    f.pack(fill="both", expand=True)
    
    ttk.Label(f, text="Register New Consumer Profile", style="Header.TLabel", foreground="#1a3a6b").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 15))
    
    fields = {}
    
    ttk.Label(f, text="CIN Number:*").grid(row=1, column=0, sticky="w", pady=5)
    fields["cin_no"] = ttk.Entry(f, width=30)
    fields["cin_no"].grid(row=1, column=1, sticky="w", pady=5)
    
    ttk.Label(f, text="Full Name:*").grid(row=2, column=0, sticky="w", pady=5)
    fields["name"] = ttk.Entry(f, width=30)
    fields["name"].grid(row=2, column=1, sticky="w", pady=5)
    
    ttk.Label(f, text="Zone:*").grid(row=3, column=0, sticky="w", pady=5)
    fields["zone"] = ttk.Combobox(f, values=[str(z) for z in utils.ZONE_RANGE], state="readonly", width=10)
    fields["zone"].grid(row=3, column=1, sticky="w", pady=5)
    
    ttk.Label(f, text="Category:*").grid(row=4, column=0, sticky="w", pady=5)
    fields["category"] = ttk.Combobox(f, values=utils.CATEGORY_OPTIONS, state="readonly", width=15)
    fields["category"].grid(row=4, column=1, sticky="w", pady=5)
    
    ttk.Label(f, text="Meter Size:*").grid(row=5, column=0, sticky="w", pady=5)
    fields["meter_size"] = ttk.Combobox(f, values=utils.METER_SIZE_OPTIONS, state="readonly", width=15)
    fields["meter_size"].grid(row=5, column=1, sticky="w", pady=5)
    
    ttk.Label(f, text="Meter Serial No:").grid(row=6, column=0, sticky="w", pady=5)
    fields["meter_serial_no"] = ttk.Entry(f, width=30)
    fields["meter_serial_no"].grid(row=6, column=1, sticky="w", pady=5)
    
    ttk.Label(f, text="Initial Meter Reading (KL):").grid(row=7, column=0, sticky="w", pady=5)
    fields["initial_meter_reading"] = ttk.Entry(f, width=20)
    fields["initial_meter_reading"].grid(row=7, column=1, sticky="w", pady=5)
    fields["initial_meter_reading"].insert(0, "0.0")
    
    ttk.Label(f, text="Contact Number:").grid(row=8, column=0, sticky="w", pady=5)
    fields["contact_number"] = ttk.Entry(f, width=30)
    fields["contact_number"].grid(row=8, column=1, sticky="w", pady=5)
    
    ttk.Label(f, text="Aadhaar/PHED No:").grid(row=9, column=0, sticky="w", pady=5)
    fields["aadhaar_phed_no"] = ttk.Entry(f, width=30)
    fields["aadhaar_phed_no"].grid(row=9, column=1, sticky="w", pady=5)
    
    ttk.Label(f, text="APL / BPL Status:").grid(row=10, column=0, sticky="w", pady=5)
    fields["apl_bpl"] = ttk.Combobox(f, values=utils.APL_BPL_OPTIONS, state="readonly", width=10)
    fields["apl_bpl"].grid(row=10, column=1, sticky="w", pady=5)
    fields["apl_bpl"].set("APL")
    
    ttk.Label(f, text="GPS Latitude:").grid(row=11, column=0, sticky="w", pady=5)
    fields["address_latitude"] = ttk.Entry(f, width=20)
    fields["address_latitude"].grid(row=11, column=1, sticky="w", pady=5)
    
    ttk.Label(f, text="GPS Longitude:").grid(row=12, column=0, sticky="w", pady=5)
    fields["address_longitude"] = ttk.Entry(f, width=20)
    fields["address_longitude"].grid(row=12, column=1, sticky="w", pady=5)
    
    ttk.Label(f, text="PIN Code:").grid(row=13, column=0, sticky="w", pady=5)
    fields["address_pin_code"] = ttk.Entry(f, width=20)
    fields["address_pin_code"].grid(row=13, column=1, sticky="w", pady=5)
    
    ttk.Label(f, text="Area/Location:").grid(row=14, column=0, sticky="w", pady=5)
    fields["address_area_location"] = ttk.Entry(f, width=40)
    fields["address_area_location"].grid(row=14, column=1, sticky="w", pady=5)
    
    ttk.Label(f, text="Landmark:").grid(row=15, column=0, sticky="w", pady=5)
    fields["address_landmark"] = ttk.Entry(f, width=40)
    fields["address_landmark"].grid(row=15, column=1, sticky="w", pady=5)
    
    ttk.Label(f, text="Consumer Status:").grid(row=16, column=0, sticky="w", pady=5)
    fields["consumer_status"] = ttk.Combobox(f, values=utils.CONSUMER_STATUS_OPTIONS, state="readonly", width=15)
    fields["consumer_status"].grid(row=16, column=1, sticky="w", pady=5)
    fields["consumer_status"].set("Active")

    # Custom Attributes Table inside Add Tab
    ttk.Label(f, text="Custom Attributes:", font=("Segoe UI", 10, "bold")).grid(row=17, column=0, sticky="nw", pady=10)
    attr_frame = ttk.Frame(f)
    attr_frame.grid(row=17, column=1, sticky="w", pady=10)
    
    attr_tree = ttk.Treeview(attr_frame, columns=("key", "value"), show="headings", height=4)
    attr_tree.heading("key", text="Attribute Name")
    attr_tree.heading("value", text="Value")
    attr_tree.column("key", width=120)
    attr_tree.column("value", width=180)
    attr_tree.grid(row=0, column=0, rowspan=2, padx=5)
    
    custom_attrs = {}
    
    def add_attr_action():
        key_win = tk.Toplevel(tab)
        key_win.title("Add Custom Attribute")
        key_win.geometry("300x180")
        key_win.grab_set()
        key_win.resizable(False, False)
        
        ttk.Label(key_win, text="Attribute Name:").pack(pady=5)
        k_ent = ttk.Entry(key_win, width=25)
        k_ent.pack(pady=2)
        
        ttk.Label(key_win, text="Attribute Value:").pack(pady=5)
        v_ent = ttk.Entry(key_win, width=25)
        v_ent.pack(pady=2)
        
        def save():
            k = k_ent.get().strip()
            v = v_ent.get().strip()
            if not k or not v:
                messagebox.showerror("Error", "Fields cannot be empty.", parent=key_win)
                return
            custom_attrs[k] = v
            attr_tree.insert("", "end", values=(k, v))
            key_win.destroy()
            
        ttk.Button(key_win, text="Add", command=save).pack(pady=10)
        
    def remove_attr_action():
        sel = attr_tree.selection()
        if not sel:
            return
        item_vals = attr_tree.item(sel[0], "values")
        key = item_vals[0]
        if key in custom_attrs:
            del custom_attrs[key]
        attr_tree.delete(sel[0])

    ttk.Button(attr_frame, text="➕ Add Row", command=add_attr_action).grid(row=0, column=1, sticky="w", padx=5, pady=2)
    ttk.Button(attr_frame, text="➖ Remove Row", command=remove_attr_action).grid(row=1, column=1, sticky="w", padx=5, pady=2)
    
    # Save Action
    def save_consumer():
        cin = fields["cin_no"].get().strip()
        name = fields["name"].get().strip()
        zone = fields["zone"].get()
        cat = fields["category"].get()
        sz = fields["meter_size"].get()
        
        # Validation
        if not cin or not name or not zone or not cat or not sz:
            messagebox.showerror("Validation Error", "Please fill all required (*) fields.", parent=tab)
            return
            
        lat_str = fields["address_latitude"].get().strip()
        long_str = fields["address_longitude"].get().strip()
        pin_str = fields["address_pin_code"].get().strip()
        
        # Build payload
        payload = {
            "cin_no": cin,
            "name": name,
            "zone": int(zone),
            "category": cat,
            "meter_size": sz,
            "meter_serial_no": fields["meter_serial_no"].get().strip(),
            "initial_meter_reading": float(fields["initial_meter_reading"].get() or 0.0),
            "contact_number": fields["contact_number"].get().strip() or None,
            "aadhaar_phed_no": fields["aadhaar_phed_no"].get().strip() or None,
            "apl_bpl": fields["apl_bpl"].get(),
            "address_latitude": float(lat_str) if lat_str else None,
            "address_longitude": float(long_str) if long_str else None,
            "address_pin_code": int(pin_str) if pin_str else None,
            "address_area_location": fields["address_area_location"].get().strip() or None,
            "address_landmark": fields["address_landmark"].get().strip() or None,
            "consumer_status": fields["consumer_status"].get(),
            "custom_attributes": custom_attrs,
            "credit_balance": 0.0,
            "outstanding_balance": 0.0,
            "is_active": True
        }
        
        def save():
            return fc.create_consumer(payload, admin["name"])
            
        def success(cin_id):
            messagebox.showinfo("Success", f"Consumer Profile {cin_id} created successfully.", parent=tab)
            # Clear form
            fields["cin_no"].delete(0, "end")
            fields["name"].delete(0, "end")
            fields["meter_serial_no"].delete(0, "end")
            fields["initial_meter_reading"].delete(0, "end")
            fields["initial_meter_reading"].insert(0, "0.0")
            fields["contact_number"].delete(0, "end")
            fields["aadhaar_phed_no"].delete(0, "end")
            fields["address_latitude"].delete(0, "end")
            fields["address_longitude"].delete(0, "end")
            fields["address_pin_code"].delete(0, "end")
            fields["address_area_location"].delete(0, "end")
            fields["address_landmark"].delete(0, "end")
            
            # Reset combs
            fields["zone"].set("")
            fields["category"].set("")
            fields["meter_size"].set("")
            fields["apl_bpl"].set("APL")
            fields["consumer_status"].set("Active")
            
            # Clear attrs
            custom_attrs.clear()
            for r in attr_tree.get_children():
                attr_tree.delete(r)
                
            # Jump to Search Tab
            notebook.select(0)
            
        def fail(err):
            messagebox.showerror("Error", f"Failed to save consumer profile:\n{err}", parent=tab)
            
        utils.run_in_thread(save, callback=success, error_callback=fail, widget=tab)

    ttk.Button(f, text="💾 Save Profile", command=save_consumer).grid(row=18, column=1, sticky="w", pady=20)

def show_edit_consumer_dialog(c, fc, utils, admin, parent_popup, success_callback):
    edit_win = tk.Toplevel(parent_popup)
    edit_win.title(f"Edit Consumer — {c['cin_no']}")
    edit_win.geometry("500x550")
    edit_win.grab_set()
    
    # Simple form inside scrollable canvas or just standard frame
    f = ttk.Frame(edit_win, padding=15)
    f.pack(fill="both", expand=True)
    
    ttk.Label(f, text=f"Update Profile for {c['cin_no']}", style="Header.TLabel", foreground="#1a3a6b").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 15))
    
    fields = {}
    
    ttk.Label(f, text="Full Name:*").grid(row=1, column=0, sticky="w", pady=4)
    fields["name"] = ttk.Entry(f, width=30)
    fields["name"].grid(row=1, column=1, sticky="w", pady=4)
    fields["name"].insert(0, c["name"])
    
    ttk.Label(f, text="Zone:*").grid(row=2, column=0, sticky="w", pady=4)
    fields["zone"] = ttk.Combobox(f, values=[str(z) for z in utils.ZONE_RANGE], state="readonly", width=10)
    fields["zone"].grid(row=2, column=1, sticky="w", pady=4)
    fields["zone"].set(str(c["zone"]))
    
    ttk.Label(f, text="Category:*").grid(row=3, column=0, sticky="w", pady=4)
    fields["category"] = ttk.Combobox(f, values=utils.CATEGORY_OPTIONS, state="readonly", width=15)
    fields["category"].grid(row=3, column=1, sticky="w", pady=4)
    fields["category"].set(c["category"])
    
    ttk.Label(f, text="Meter Size:*").grid(row=4, column=0, sticky="w", pady=4)
    fields["meter_size"] = ttk.Combobox(f, values=utils.METER_SIZE_OPTIONS, state="readonly", width=15)
    fields["meter_size"].grid(row=4, column=1, sticky="w", pady=4)
    fields["meter_size"].set(c["meter_size"])
    
    ttk.Label(f, text="Meter Serial No:").grid(row=5, column=0, sticky="w", pady=4)
    fields["meter_serial_no"] = ttk.Entry(f, width=30)
    fields["meter_serial_no"].grid(row=5, column=1, sticky="w", pady=4)
    fields["meter_serial_no"].insert(0, c.get("meter_serial_no", ""))
    
    ttk.Label(f, text="Contact Number:").grid(row=6, column=0, sticky="w", pady=4)
    fields["contact_number"] = ttk.Entry(f, width=30)
    fields["contact_number"].grid(row=6, column=1, sticky="w", pady=4)
    fields["contact_number"].insert(0, c.get("contact_number") or "")
    
    ttk.Label(f, text="Aadhaar/PHED No:").grid(row=7, column=0, sticky="w", pady=4)
    fields["aadhaar_phed_no"] = ttk.Entry(f, width=30)
    fields["aadhaar_phed_no"].grid(row=7, column=1, sticky="w", pady=4)
    fields["aadhaar_phed_no"].insert(0, c.get("aadhaar_phed_no") or "")
    
    ttk.Label(f, text="APL / BPL Status:").grid(row=8, column=0, sticky="w", pady=4)
    fields["apl_bpl"] = ttk.Combobox(f, values=utils.APL_BPL_OPTIONS, state="readonly", width=10)
    fields["apl_bpl"].grid(row=8, column=1, sticky="w", pady=4)
    fields["apl_bpl"].set(c.get("apl_bpl", "APL"))
    
    ttk.Label(f, text="GPS Latitude:").grid(row=9, column=0, sticky="w", pady=4)
    fields["address_latitude"] = ttk.Entry(f, width=20)
    fields["address_latitude"].grid(row=9, column=1, sticky="w", pady=4)
    if c.get("address_latitude") is not None:
        fields["address_latitude"].insert(0, str(c["address_latitude"]))
        
    ttk.Label(f, text="GPS Longitude:").grid(row=10, column=0, sticky="w", pady=4)
    fields["address_longitude"] = ttk.Entry(f, width=20)
    fields["address_longitude"].grid(row=10, column=1, sticky="w", pady=4)
    if c.get("address_longitude") is not None:
        fields["address_longitude"].insert(0, str(c["address_longitude"]))
        
    ttk.Label(f, text="PIN Code:").grid(row=11, column=0, sticky="w", pady=4)
    fields["address_pin_code"] = ttk.Entry(f, width=20)
    fields["address_pin_code"].grid(row=11, column=1, sticky="w", pady=4)
    if c.get("address_pin_code") is not None:
        fields["address_pin_code"].insert(0, str(c["address_pin_code"]))
        
    ttk.Label(f, text="Area/Location:").grid(row=12, column=0, sticky="w", pady=4)
    fields["address_area_location"] = ttk.Entry(f, width=40)
    fields["address_area_location"].grid(row=12, column=1, sticky="w", pady=4)
    fields["address_area_location"].insert(0, c.get("address_area_location", "") or "")
    
    ttk.Label(f, text="Landmark:").grid(row=13, column=0, sticky="w", pady=4)
    fields["address_landmark"] = ttk.Entry(f, width=40)
    fields["address_landmark"].grid(row=13, column=1, sticky="w", pady=4)
    fields["address_landmark"].insert(0, c.get("address_landmark", "") or "")

    def save_edits():
        name = fields["name"].get().strip()
        zone = fields["zone"].get()
        cat = fields["category"].get()
        sz = fields["meter_size"].get()
        
        if not name or not zone or not cat or not sz:
            messagebox.showerror("Error", "Required fields cannot be empty.", parent=edit_win)
            return
            
        lat_str = fields["address_latitude"].get().strip()
        long_str = fields["address_longitude"].get().strip()
        pin_str = fields["address_pin_code"].get().strip()
        
        payload = {
            "name": name,
            "zone": int(zone),
            "category": cat,
            "meter_size": sz,
            "meter_serial_no": fields["meter_serial_no"].get().strip(),
            "contact_number": fields["contact_number"].get().strip() or None,
            "aadhaar_phed_no": fields["aadhaar_phed_no"].get().strip() or None,
            "apl_bpl": fields["apl_bpl"].get(),
            "address_latitude": float(lat_str) if lat_str else None,
            "address_longitude": float(long_str) if long_str else None,
            "address_pin_code": int(pin_str) if pin_str else None,
            "address_area_location": fields["address_area_location"].get().strip() or None,
            "address_landmark": fields["address_landmark"].get().strip() or None
        }
        
        def save():
            fc.update_consumer(c["cin_no"], payload, admin["name"])
            
        def done(res):
            messagebox.showinfo("Success", "Consumer profile updated successfully.", parent=edit_win)
            edit_win.destroy()
            success_callback()
            
        def fail(err):
            messagebox.showerror("Error", f"Failed to update consumer:\n{err}", parent=edit_win)
            
        utils.run_in_thread(save, callback=done, error_callback=fail, widget=edit_win)

    ttk.Button(f, text="💾 Save Changes", command=save_edits).grid(row=14, column=1, sticky="w", pady=15)

# ----------------------------------------------------
# TAB 3: Bulk Import CSD
# ----------------------------------------------------
def build_import_tab(tab, fc, utils, be, admin):
    ttk.Label(tab, text="Bulk Import Consumer Static Data (CSD) from Excel", style="Header.TLabel", foreground="#1a3a6b").pack(anchor="w", pady=(0, 10))
    
    desc = ("Upload multiple consumer profiles at once using our spreadsheet template.\n"
            "1. Download the Template.\n"
            "2. Fill in details (cin_no, name, category, initial reading, zone, etc.).\n"
            "3. Select the file and click 'Import'.")
    ttk.Label(tab, text=desc).pack(anchor="w", pady=(0, 15))
    
    # Download Template Button
    def download_tmpl():
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Save Consumer Template")
        if not path:
            return
        try:
            xlsx_bytes = utils.get_excel_template_consumers()
            with open(path, "wb") as f:
                f.write(xlsx_bytes)
            messagebox.showinfo("Success", f"Template saved successfully at:\n{path}", parent=tab)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save template:\n{e}", parent=tab)
            
    ttk.Button(tab, text="📥 Download Excel Template", command=download_tmpl).pack(anchor="w", pady=5)
    
    # Import Frame
    imp_frame = ttk.LabelFrame(tab, text="Select & Run Import", padding=15)
    imp_frame.pack(fill="x", pady=15)
    
    file_path_var = tk.StringVar()
    ttk.Label(imp_frame, text="File Path:").grid(row=0, column=0, sticky="w")
    ttk.Entry(imp_frame, textvariable=file_path_var, width=50, state="readonly").grid(row=0, column=1, padx=10)
    
    def pick_file():
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")], title="Select Consumers Data File")
        if path:
            file_path_var.set(path)
            
    ttk.Button(imp_frame, text="📂 Browse...", command=pick_file).grid(row=0, column=2)
    
    prog_bar = ttk.Progressbar(imp_frame, orient="horizontal", mode="indeterminate")
    status_lbl = ttk.Label(imp_frame, text="", font=("Segoe UI", 9, "italic"))
    
    def start_import():
        path = file_path_var.get()
        if not path:
            messagebox.showerror("Error", "Please select a file to import.", parent=tab)
            return
            
        prog_bar.grid(row=2, column=0, columnspan=3, sticky="ew", pady=10)
        status_lbl.grid(row=3, column=0, columnspan=3, sticky="w")
        prog_bar.start(10)
        status_lbl.config(text="Parsing spreadsheet contents...")
        
        def run():
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                raise ValueError("The excel sheet is empty or contains no data rows.")
                
            # Maps headers
            headers = [str(cell).strip() for cell in rows[0]]
            data_list = []
            
            for r_idx, row in enumerate(rows[1:]):
                if not row or not any(row):
                    continue # skip blank lines
                # Build dict
                payload = {}
                for h_idx, h in enumerate(headers):
                    if h_idx < len(row):
                        payload[h] = row[h_idx]
                
                # Check for formatting and data cleanliness
                if not payload.get("cin_no"):
                    raise ValueError(f"Missing required 'cin_no' on row {r_idx + 2}")
                if not payload.get("name"):
                    raise ValueError(f"Missing required 'name' on row {r_idx + 2}")
                if payload.get("zone") is None:
                    raise ValueError(f"Missing required 'zone' on row {r_idx + 2}")
                if not payload.get("category"):
                    raise ValueError(f"Missing required 'category' on row {r_idx + 2}")
                if not payload.get("meter_size"):
                    raise ValueError(f"Missing required 'meter_size' on row {r_idx + 2}")
                    
                data_list.append(payload)
                
            status_lbl.config(text=f"Importing {len(data_list)} consumer records to Firebase...")
            res = fc.bulk_create_consumers(data_list, admin["name"])
            return res
            
        def done(res):
            prog_bar.stop()
            prog_bar.grid_remove()
            status_lbl.config(text="")
            
            success = res["success"]
            errors = res["errors"]
            
            msg = f"Bulk import complete!\n\nSuccessfully Imported: {success} profiles"
            if errors:
                msg += f"\nFailed Rows: {len(errors)}\n\nFirst few errors:\n"
                for err in errors[:5]:
                    msg += f"• Row {err['row']}: {err['error']}\n"
                messagebox.showwarning("Import Warning", msg, parent=tab)
            else:
                messagebox.showinfo("Success", msg, parent=tab)
                
            file_path_var.set("")
            
        def fail(err):
            prog_bar.stop()
            prog_bar.grid_remove()
            status_lbl.config(text="")
            messagebox.showerror("Error", f"Failed running bulk import:\n{err}", parent=tab)
            
        utils.run_in_thread(run, callback=done, error_callback=fail, widget=tab)

    ttk.Button(imp_frame, text="🚀 Import Consumers", command=start_import).grid(row=1, column=1, sticky="w", pady=15)

# ----------------------------------------------------
# TAB 4: Export CSD
# ----------------------------------------------------
def build_export_tab(tab, fc, utils, be, admin):
    ttk.Label(tab, text="Export Consumer Static Data (CSD) to Excel", style="Header.TLabel", foreground="#1a3a6b").pack(anchor="w", pady=(0, 10))
    ttk.Label(tab, text="Filter your query before generating the spreadsheet:").pack(anchor="w", pady=(0, 15))
    
    f_frame = ttk.LabelFrame(tab, text="Export Settings", padding=15)
    f_frame.pack(fill="x", pady=5)
    
    ttk.Label(f_frame, text="Zone Filter:").grid(row=0, column=0, sticky="w", pady=5)
    zone_var = tk.StringVar(value="All")
    ttk.Combobox(f_frame, textvariable=zone_var, values=["All"] + [str(z) for z in utils.ZONE_RANGE], state="readonly", width=10).grid(row=0, column=1, sticky="w", pady=5, padx=10)
    
    ttk.Label(f_frame, text="Status Filter:").grid(row=1, column=0, sticky="w", pady=5)
    status_var = tk.StringVar(value="All")
    ttk.Combobox(f_frame, textvariable=status_var, values=["All"] + utils.CONSUMER_STATUS_OPTIONS, state="readonly", width=15).grid(row=1, column=1, sticky="w", pady=5, padx=10)
    
    def run_export():
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Save Exported Consumers")
        if not path:
            return
            
        z_filter = zone_var.get()
        s_filter = status_var.get()
        
        def run():
            filters = {}
            if z_filter != "All":
                filters["zone"] = int(z_filter)
            if s_filter != "All":
                filters["status"] = s_filter
                
            consumers = fc.list_consumers(filters)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Consumers CSD"
            
            headers = [
                "cin_no", "name", "zone", "contact_number", "category", "meter_size", 
                "meter_serial_no", "initial_meter_reading", "last_reading", "address_longitude", 
                "address_latitude", "address_pin_code", "address_area_location", 
                "address_landmark", "aadhaar_phed_no", "apl_bpl", "consumer_status", 
                "outstanding_balance", "credit_balance"
            ]
            
            ws.append(headers)
            # Styling headers
            header_font = openpyxl.styles.Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            header_fill = openpyxl.styles.PatternFill(start_color="1A3A6B", end_color="1A3A6B", fill_type="solid")
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                
            for c in consumers:
                ws.append([
                    c.get("cin_no"),
                    c.get("name"),
                    c.get("zone"),
                    c.get("contact_number"),
                    c.get("category"),
                    c.get("meter_size"),
                    c.get("meter_serial_no"),
                    c.get("initial_meter_reading"),
                    c.get("last_reading"),
                    c.get("address_longitude"),
                    c.get("address_latitude"),
                    c.get("address_pin_code"),
                    c.get("address_area_location"),
                    c.get("address_landmark"),
                    c.get("aadhaar_phed_no"),
                    c.get("apl_bpl"),
                    c.get("consumer_status"),
                    c.get("outstanding_balance"),
                    c.get("credit_balance")
                ])
                
            for col in ws.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
                
            wb.save(path)
            
        def done(res):
            messagebox.showinfo("Success", f"CSD Excel export created at:\n{path}", parent=tab)
            
        def fail(err):
            messagebox.showerror("Error", f"Export failed:\n{err}", parent=tab)
            
        utils.run_in_thread(run, callback=done, error_callback=fail, widget=tab)

    ttk.Button(f_frame, text="📤 Export to Excel Spreadsheet", command=run_export).grid(row=2, column=1, sticky="w", pady=15, padx=10)

# ----------------------------------------------------
# TAB 5: Meter Replacement
# ----------------------------------------------------
def build_replacement_tab(tab, fc, utils, be, admin):
    ttk.Label(tab, text="Record Meter Replacement Event", style="Header.TLabel", foreground="#1a3a6b").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 15))
    
    # Left search area, Right update form
    lf = ttk.Frame(tab)
    lf.grid(row=1, column=0, sticky="n", padx=5)
    
    rf = ttk.LabelFrame(tab, text="Enter Replacement Log Details", padding=15)
    rf.grid(row=1, column=1, sticky="n", padx=15)
    
    # Search elements
    ttk.Label(lf, text="Search Consumer (CIN):").pack(anchor="w")
    cin_ent = ttk.Entry(lf, width=22)
    cin_ent.pack(anchor="w", pady=5)
    
    consumer_name_var = tk.StringVar(value="—")
    old_serial_var = tk.StringVar(value="—")
    last_reading_var = tk.StringVar(value="—")
    
    fields = {}
    
    def search_consumer():
        cin = cin_ent.get().strip()
        if not cin:
            return
            
        def fetch():
            return fc.get_consumer(cin)
            
        def done(c):
            if c:
                consumer_name_var.set(c["name"])
                old_serial_var.set(c.get("meter_serial_no", "None"))
                last_reading_var.set(f"{c.get('last_reading', 0.0):.2f} KL")
                
                # Pre-fill replacement defaults
                fields["new_serial"].focus()
            else:
                messagebox.showerror("Error", f"No consumer found with CIN: {cin}", parent=tab)
                consumer_name_var.set("—")
                old_serial_var.set("—")
                last_reading_var.set("—")
                
        utils.run_in_thread(fetch, callback=done, widget=tab)
        
    ttk.Button(lf, text="🔍 Search Profile", command=search_consumer).pack(anchor="w", pady=5)
    
    # Display info
    info_box = ttk.LabelFrame(lf, text="Target Consumer", padding=10)
    info_box.pack(anchor="w", fill="x", pady=10)
    
    ttk.Label(info_box, text="Name:").grid(row=0, column=0, sticky="w", pady=2)
    ttk.Label(info_box, textvariable=consumer_name_var, font=("Segoe UI", 9, "bold")).grid(row=0, column=1, sticky="w", pady=2, padx=5)
    
    ttk.Label(info_box, text="Old Serial:").grid(row=1, column=0, sticky="w", pady=2)
    ttk.Label(info_box, textvariable=old_serial_var).grid(row=1, column=1, sticky="w", pady=2, padx=5)
    
    ttk.Label(info_box, text="Last Reading:").grid(row=2, column=0, sticky="w", pady=2)
    ttk.Label(info_box, textvariable=last_reading_var).grid(row=2, column=1, sticky="w", pady=2, padx=5)
    
    # Form details
    ttk.Label(rf, text="New Meter Serial No:*").grid(row=0, column=0, sticky="w", pady=5)
    fields["new_serial"] = ttk.Entry(rf, width=25)
    fields["new_serial"].grid(row=0, column=1, sticky="w", pady=5)
    
    ttk.Label(rf, text="Replacement Date:*").grid(row=1, column=0, sticky="w", pady=5)
    fields["rep_date"] = ttk.Entry(rf, width=20)
    fields["rep_date"].grid(row=1, column=1, sticky="w", pady=5)
    fields["rep_date"].insert(0, datetime.now().strftime("%d-%m-%Y"))
    
    ttk.Label(rf, text="New Initial Reading (KL):*").grid(row=2, column=0, sticky="w", pady=5)
    fields["new_reading"] = ttk.Entry(rf, width=15)
    fields["new_reading"].grid(row=2, column=1, sticky="w", pady=5)
    fields["new_reading"].insert(0, "0.0")
    
    def save_replacement():
        cin = cin_ent.get().strip()
        new_s = fields["new_serial"].get().strip()
        rep_d = fields["rep_date"].get().strip()
        new_r = fields["new_reading"].get().strip()
        
        if not cin or consumer_name_var.get() == "—":
            messagebox.showerror("Error", "Please select a valid consumer profile first.", parent=tab)
            return
        if not new_s or not rep_d or not new_r:
            messagebox.showerror("Error", "Please fill in all replacement form fields.", parent=tab)
            return
            
        try:
            r_val = float(new_r)
        except ValueError:
            messagebox.showerror("Error", "New initial reading must be a valid float value.", parent=tab)
            return
            
        confirm = messagebox.askyesno("Confirm Replacement", f"This will reset the consumer's meter details to Serial: {new_s} and Reading: {new_r} KL. Outstanding balance remains unchanged. Confirm?", parent=tab)
        if not confirm:
            return
            
        def save():
            return fc.record_meter_replacement(
                cin_no=cin,
                old_serial=old_serial_var.get(),
                new_serial=new_s,
                replacement_date=rep_d,
                new_initial_reading_kl=r_val,
                admin_name=admin["name"]
            )
            
        def done(log_id):
            messagebox.showinfo("Success", f"Meter replacement log written under ID:\n{log_id}", parent=tab)
            # Clear
            cin_ent.delete(0, "end")
            fields["new_serial"].delete(0, "end")
            fields["new_reading"].delete(0, "end")
            fields["new_reading"].insert(0, "0.0")
            
            consumer_name_var.set("—")
            old_serial_var.set("—")
            last_reading_var.set("—")
            
        def fail(err):
            messagebox.showerror("Error", f"Failed saving meter replacement:\n{err}", parent=tab)
            
        utils.run_in_thread(save, callback=done, error_callback=fail, widget=tab)

    ttk.Button(rf, text="⚙️ Record Meter Replacement", command=save_replacement).grid(row=3, column=1, sticky="w", pady=15)
