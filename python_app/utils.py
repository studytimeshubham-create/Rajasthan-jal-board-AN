import os
import sys
import subprocess
import random
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from weasyprint import HTML

# Design Tokens (Claude Editorial Style)
UI_COLORS = {
    "primary": "#cc785c",
    "primary_active": "#a9583e",
    "primary_disabled": "#e6dfd8",
    "ink": "#141413",
    "body": "#3d3d3a",
    "body_strong": "#252523",
    "muted": "#6c6a64",
    "muted_soft": "#8e8b82",
    "hairline": "#e6dfd8",
    "hairline_soft": "#ebe6df",
    "canvas": "#faf9f5",
    "surface_soft": "#f5f0e8",
    "surface_card": "#efe9de",
    "surface_cream_strong": "#e8e0d2",
    "surface_dark": "#181715",
    "surface_dark_elevated": "#252320",
    "surface_dark_soft": "#1f1e1b",
    "on_primary": "#ffffff",
    "on_dark": "#faf9f5",
    "on_dark_soft": "#a09d96",
    "accent_teal": "#5db8a6",
    "accent_amber": "#e8a55a",
    "success": "#5db872",
    "warning": "#d4a017",
    "error": "#c64545",
}

# Constants
CONSUMER_STATUS_OPTIONS = ["Active", "Inactive", "Meter Faulty", "Disconnected", "Disputed"]
METER_SIZE_OPTIONS = ["15mm", "20mm", "25mm", "40mm", "50mm", "80mm", "100mm", "150mm"]
CATEGORY_OPTIONS = ["Domestic", "Non-Domestic", "Industrial"]
APL_BPL_OPTIONS = ["APL", "BPL"]
ZONE_RANGE = range(1, 21)  # zones 1–20

READER_ROLE_OPTIONS = ["Reader", "Cashier"]
SUPPLY_TYPE_OPTIONS = ["PHED", "Own", "Mixed"]
SEWERAGE_SUB_CATEGORY_OPTIONS = [
    "Hotel",
    "Restaurant",
    "Cinema",
    "Car/Truck Service Station",
    "Scooter Service Station",
    "Other Industrial/Commercial",
    "Domestic",
    "House > 200sqm"
]

def format_currency(amount: float) -> str:
    """Formats amount to Indian Rupees format, e.g. ₹1,234.56."""
    try:
        if amount is None:
            amount = 0.0
        # Basic formatting for Indian Rupees representation
        # Indian standard separates first thousands, then hundreds
        s = f"{amount:.2f}"
        parts = s.split(".")
        num = parts[0]
        dec = parts[1]
        
        if len(num) > 3:
            last_three = num[-3:]
            remaining = num[:-3]
            # Group by 2 digits
            groups = []
            while remaining:
                if len(remaining) >= 2:
                    groups.append(remaining[-2:])
                    remaining = remaining[:-2]
                else:
                    groups.append(remaining)
                    remaining = ""
            groups.reverse()
            formatted_num = ",".join(groups) + "," + last_three
        else:
            formatted_num = num
            
        return f"₹{formatted_num}.{dec}"
    except Exception:
        return f"₹{amount:.2f}"

def format_kl(value: float) -> str:
    """Formats a consumption value as '12.34 KL'."""
    val = value if value is not None else 0.0
    return f"{val:.2f} KL"

def format_date(dt) -> str:
    """Accepts datetime, date, or Firestore Timestamp and returns 'DD-MM-YYYY'."""
    if dt is None:
        return ""
    if hasattr(dt, "strftime"):
        return dt.strftime("%d-%m-%Y")
    # If it is a Firestore timestamp (which may have datetime attribute or be string)
    if hasattr(dt, "datetime"):
        return dt.datetime.strftime("%d-%m-%Y")
    # If it's a string, attempt conversion
    s = str(dt)
    if "T" in s:
        # ISO format
        try:
            return datetime.fromisoformat(s).strftime("%d-%m-%Y")
        except ValueError:
            pass
    # Otherwise return as-is
    return s

def parse_date(s: str) -> date:
    """Parses 'DD-MM-YYYY' string into a Python date object."""
    if not s:
        return date.today()
    try:
        return datetime.strptime(s.strip(), "%d-%m-%Y").date()
    except Exception:
        # Fallback to ISO parsing if DD-MM-YYYY fails
        try:
            return datetime.fromisoformat(s.strip()).date()
        except Exception:
            return date.today()

def today_str() -> str:
    """Returns today's date formatted as 'DD-MM-YYYY'."""
    return date.today().strftime("%d-%m-%Y")

def add_months(d: date, months: int) -> date:
    """Adds N calendar months using dateutil.relativedelta.
    For example, March 31 + 1 month -> April 30.
    """
    return d + relativedelta(months=months)

def run_in_thread(fn, *args, callback=None, error_callback=None, widget=None, **kwargs):
    """Runs a function in a background daemon thread.
    Coordinates callbacks back to the Tkinter thread safely via widget.after.
    """
    import threading
    import traceback

    def widget_exists():
        if not widget:
            return False
        try:
            return bool(widget.winfo_exists())
        except Exception:
            return False

    def is_destroyed_widget_error(exc):
        return exc.__class__.__name__ == "TclError" and "invalid command name" in str(exc)

    def dispatch(cb, value):
        try:
            cb(value)
        except Exception as callback_error:
            if not is_destroyed_widget_error(callback_error):
                traceback.print_exc()

    def schedule(cb, value):
        if not cb:
            return
        if widget:
            try:
                if not widget.winfo_exists():
                    return
                widget.after(0, lambda value=value: dispatch(cb, value))
            except Exception as schedule_error:
                if not is_destroyed_widget_error(schedule_error):
                    traceback.print_exc()
        else:
            dispatch(cb, value)

    def wrapper():
        try:
            result = fn(*args, **kwargs)
            schedule(callback, result)
        except Exception as e:
            traceback.print_exc()
            schedule(error_callback, e)

    thread = threading.Thread(target=wrapper)
    thread.daemon = True
    thread.start()
    return thread

def generate_receipt_number() -> str:
    """Generates unique receipt numbers: RJB-YYYYMMDD-XXXXXX (6-digit zero-padded random)."""
    date_part = datetime.now().strftime("%Y%m%d")
    random_part = f"{random.randint(0, 999999):06d}"
    return f"RJB-{date_part}-{random_part}"

def load_pdf_template(template_name: str) -> str:
    """Loads HTML template from pdf_templates/{template_name}.html."""
    # Find paths relative to utils location or parent directory
    script_dir = os.path.dirname(os.path.abspath(__file__))
    path = os.path.join(script_dir, "pdf_templates", f"{template_name}.html")
    if not os.path.exists(path):
        # Check parent folder
        parent_dir = os.path.dirname(script_dir)
        path = os.path.join(parent_dir, "pdf_templates", f"{template_name}.html")
        if not os.path.exists(path):
            raise FileNotFoundError(f"Template '{template_name}' not found at: {path}")
            
    with open(path, "r", encoding="utf-8") as f:
        return f.read()

def render_pdf_to_bytes(html_string: str) -> bytes:
    """Converts an HTML string to PDF bytes using WeasyPrint."""
    return HTML(string=html_string).write_pdf()

def render_pdf_to_file(html_string: str, output_path: str) -> None:
    """Converts an HTML string and saves it directly to a file using WeasyPrint."""
    HTML(string=html_string).write_pdf(target=output_path)

def open_pdf(path: str) -> None:
    """Opens a PDF file in the system default reader program."""
    try:
        if sys.platform.startswith("win"):
            os.startfile(path)
        elif sys.platform.startswith("darwin"):
            subprocess.call(["open", path])
        else:
            subprocess.call(["xdg-open", path])
    except Exception as e:
        print(f"Error opening PDF: {e}")

# ----------------------------------------------------
# Excel Templates Generators
# ----------------------------------------------------
def _style_excel_header(ws, headers):
    ws.append(headers)
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1A3A6B", end_color="1A3A6B", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 25

def get_excel_template_consumers() -> bytes:
    """Returns openpyxl Excel bytes with headers and a demo row for Consumer bulk upload."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Create Consumers"
    
    headers = [
        "cin_no", "name", "zone", "contact_number", "category", "meter_size", 
        "meter_serial_no", "initial_meter_reading", "address_longitude", 
        "address_latitude", "address_pin_code", "address_area_location", 
        "address_landmark", "aadhaar_phed_no", "apl_bpl", "consumer_status"
    ]
    _style_excel_header(ws, headers)
    
    demo_row = [
        "RJB-0000001", "Shubham Gupta", 1, 9876543210, "Domestic", "15mm",
        "MS-98218A", 150.5, 75.806, 26.915, 302001, "Malviya Nagar",
        "Near PHED Office", "1234-5678-9012", "APL", "Active"
    ]
    ws.append(demo_row)
    
    # Auto-fit columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def get_excel_template_payments() -> bytes:
    """Returns openpyxl Excel bytes with headers and a demo row for Payments bulk upload."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Record Payments"
    
    headers = ["cin_no", "amount", "payment_mode", "emitra_key", "payment_date", "notes"]
    _style_excel_header(ws, headers)
    
    demo_row = ["RJB-0000001", 1250, "Cash", "", today_str(), "Monthly regular payment"]
    ws.append(demo_row)
    
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def get_excel_template_readings() -> bytes:
    """Returns openpyxl Excel bytes with headers and a demo row for Readings bulk upload."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulk Edit Readings"
    
    headers = ["reading_id", "cin_no", "cycle_id", "reader_name", "previous_reading", "current_reading", "notes"]
    _style_excel_header(ws, headers)
    
    demo_row = ["READ_TEMP_1", "RJB-0000001", "BC-20250101", "Shubham Gupta", 150.5, 175.2, "Manual correction"]
    ws.append(demo_row)
    
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
