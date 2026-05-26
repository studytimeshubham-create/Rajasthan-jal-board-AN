import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from datetime import datetime, date
import os
import matplotlib
matplotlib.use("TkAgg")
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

def get_frame(parent, fc, utils, be, admin) -> ttk.Frame:
    frame = ttk.Frame(parent)
    
    notebook = ttk.Notebook(frame)
    notebook.pack(fill="both", expand=True)
    
    # 7 Tabs
    tab1 = ttk.Frame(notebook, padding=10) # Reader activity
    tab2 = ttk.Frame(notebook, padding=10) # Billing summary
    tab3 = ttk.Frame(notebook, padding=10) # Zone collections
    tab4 = ttk.Frame(notebook, padding=10) # Outstanding balance
    tab5 = ttk.Frame(notebook, padding=10) # Skipped list
    tab6 = ttk.Frame(notebook, padding=10) # Consumer ledger
    tab7 = ttk.Frame(notebook, padding=10) # Backup database
    
    notebook.add(tab1, text="👷 Reader Logs")
    notebook.add(tab2, text="📊 Billing summary")
    notebook.add(tab3, text="🗺️ Zone Collections")
    notebook.add(tab4, text="💰 Outstandings")
    notebook.add(tab5, text="❌ Skippeds")
    notebook.add(tab6, text="📖 Ledger Statements")
    notebook.add(tab7, text="💾 Database Backup")
    
    build_reader_activity_tab(tab1, fc, utils, be, admin)
    build_billing_summary_tab(tab2, fc, utils, be, admin)
    build_zone_collection_tab(tab3, fc, utils, be, admin)
    build_outstanding_tab(tab4, fc, utils, be, admin)
    build_skipped_tab(tab5, fc, utils, be, admin)
    build_ledger_statement_tab(tab6, fc, utils, be, admin)
    build_backup_tab(tab7, fc, utils, be, admin)
    
    return frame

# ----------------------------------------------------
# TAB 1: Meter Reader Activity Logs
# ----------------------------------------------------
def build_reader_activity_tab(tab, fc, utils, be, admin):
    f_frame = ttk.Frame(tab, padding=5)
    f_frame.pack(fill="x", pady=(0, 10))
    
    ttk.Label(f_frame, text="Date From:").pack(side="left", padx=3)
    from_ent = ttk.Entry(f_frame, width=10)
    from_ent.pack(side="left", padx=3)
    from_ent.insert(0, utils.today_str())
    
    ttk.Label(f_frame, text="Date To:").pack(side="left", padx=3)
    to_ent = ttk.Entry(f_frame, width=10)
    to_ent.pack(side="left", padx=3)
    to_ent.insert(0, utils.today_str())
    
    ttk.Label(f_frame, text="Reader:").pack(side="left", padx=3)
    reader_var = tk.StringVar(value="All")
    reader_cb = ttk.Combobox(f_frame, textvariable=reader_var, width=15, state="readonly")
    reader_cb.pack(side="left", padx=3)
    
    ttk.Label(f_frame, text="Zone:").pack(side="left", padx=3)
    zone_var = tk.StringVar(value="All")
    zone_cb = ttk.Combobox(f_frame, textvariable=zone_var, values=["All"] + [str(z) for z in utils.ZONE_RANGE], width=5, state="readonly")
    zone_cb.pack(side="left", padx=3)
    
    # Readers dropdown loader
    def load_readers():
        def fetch():
            return fc.list_meter_readers()
        def done(readers):
            reader_cb.config(values=["All"] + [r["name"] for r in readers])
            # map names to uid locally if needed
            tab.readers_map = {r["name"]: r["uid"] for r in readers}
        utils.run_in_thread(fetch, callback=done, widget=tab)
        
    load_readers()
    
    tree_columns = ("reader", "emp_id", "cin", "name", "date", "prev", "curr", "consumption", "status")
    tree = ttk.Treeview(tab, columns=tree_columns, show="headings")
    tree.heading("reader", text="Reader")
    tree.heading("emp_id", text="Emp ID")
    tree.heading("cin", text="CIN")
    tree.heading("name", text="Consumer Name")
    tree.heading("date", text="Date")
    tree.heading("prev", text="Prev (KL)")
    tree.heading("curr", text="Curr (KL)")
    tree.heading("consumption", text="Usage (KL)")
    tree.heading("status", text="Status")
    
    for c in tree_columns:
        tree.heading(c, text=c.capitalize())
        if c in ["prev", "curr", "consumption"]:
            tree.column(c, width=75, anchor="e")
        else:
            tree.column(c, width=100, anchor="w" if c != "cin" else "center")
            
    scrollbar = ttk.Scrollbar(tab, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    tree.pack(side="top", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    summary_lbl = ttk.Label(tab, text="Total Consumption Billed: 0.00 KL | Active Logs Count: 0", font=("Segoe UI", 9, "bold"))
    summary_lbl.pack(anchor="w", pady=5)
    
    activity_cache = []
    
    def search():
        for r in tree.get_children():
            tree.delete(r)
            
        r_name = reader_var.get()
        r_uid = tab.readers_map.get(r_name) if r_name != "All" else None
        zone_f = zone_var.get()
        z_val = int(zone_f) if zone_f != "All" else None
        
        d_from = from_ent.get().strip()
        d_to = to_ent.get().strip()
        
        def fetch():
            return fc.get_meter_reader_activity(d_from, d_to, r_uid, z_val)
            
        def done(logs):
            nonlocal activity_cache
            activity_cache = logs
            
            tot_cons = 0.0
            for l in logs:
                c_val = l.get("consumption") or 0.0
                tot_cons += c_val
                
                tree.insert("", "end", values=(
                    l.get("reader_name"),
                    l.get("employee_id"),
                    l.get("cin_no"),
                    l.get("consumer_name"),
                    l.get("reading_date"),
                    f"{l.get('previous_reading', 0.0):.2f}",
                    f"{l.get('current_reading', 0.0):.2f}" if l.get("current_reading") is not None else "",
                    f"{c_val:.2f}",
                    l.get("status")
                ))
            summary_lbl.config(text=f"Total Consumption Billed: {tot_cons:.2f} KL | Active Logs Count: {len(logs)}")
            
        utils.run_in_thread(fetch, callback=done, widget=tab)

    ttk.Button(f_frame, text="🔍 Filter Logs", command=search).pack(side="left", padx=15)
    
    def export_excel():
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Save Activity Logs")
        if not path:
            return
            
        def run():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reader Logs"
            
            headers = ["reader_name", "employee_id", "cin_no", "consumer_name", "reading_date", "previous_reading", "current_reading", "consumption", "status"]
            ws.append(headers)
            for l in activity_cache:
                ws.append([
                    l.get("reader_name"),
                    l.get("employee_id"),
                    l.get("cin_no"),
                    l.get("consumer_name"),
                    l.get("reading_date"),
                    l.get("previous_reading"),
                    l.get("current_reading"),
                    l.get("consumption"),
                    l.get("status")
                ])
            wb.save(path)
            
        def done(res):
            messagebox.showinfo("Success", "Spreadsheet saved.", parent=tab)
            
        utils.run_in_thread(run, callback=done, widget=tab)
        
    ttk.Button(f_frame, text="📤 Export Excel", command=export_excel).pack(side="right", padx=5)

# ----------------------------------------------------
# TAB 2: Consolidated Billing Summary & Matplotlib
# ----------------------------------------------------
def build_billing_summary_tab(tab, fc, utils, be, admin):
    f_frame = ttk.Frame(tab, padding=5)
    f_frame.pack(fill="x", pady=(0, 10))
    
    ttk.Label(f_frame, text="Select Billing Cycle:").pack(side="left", padx=5)
    cycle_var = tk.StringVar()
    cycle_cb = ttk.Combobox(f_frame, textvariable=cycle_var, width=18, state="readonly")
    cycle_cb.pack(side="left", padx=5)
    
    # Load cycles
    def refresh_cycles():
        def fetch():
            return fc.list_billing_cycles()
        def done(cycles):
            cycle_cb.config(values=[c["cycle_id"] for c in cycles])
            if cycles:
                cycle_cb.set(cycles[0]["cycle_id"])
        utils.run_in_thread(fetch, callback=done, widget=tab)
        
    refresh_cycles()
    
    # Financial metrics frame
    cards_frame = ttk.Frame(tab)
    cards_frame.pack(fill="x", pady=10)
    
    kpis = {}
    kpi_fields = [
        ("Total Billed Amount", "billed_val", "#1a3a6b"),
        ("Total Collections In", "collections_val", "#2a7a2a"),
        ("Deficit Outstanding", "deficit_val", "#c0392b")
    ]
    for idx, (name, key, col) in enumerate(kpi_fields):
        card = ttk.Frame(cards_frame, style="Card.TFrame", padding=10)
        card.grid(row=0, column=idx, sticky="ew", padx=10)
        cards_frame.grid_columnconfigure(idx, weight=1)
        
        ttk.Label(card, text=name, font=("Segoe UI", 9, "bold")).pack(anchor="w")
        val_lbl = ttk.Label(card, text="₹0.00", font=("Segoe UI", 16, "bold"), foreground=col)
        val_lbl.pack(anchor="w", pady=(5, 0))
        kpis[key] = val_lbl
        
    # Chart frame
    chart_frame = ttk.Frame(tab)
    chart_frame.pack(fill="both", expand=True, pady=15)
    
    canvas_container = [None]
    
    def generate_report():
        c_id = cycle_var.get()
        if not c_id:
            return
            
        def fetch():
            metrics = fc.get_billing_summary(c_id)
            zones_data = fc.get_zone_collection_report(c_id)
            return metrics, zones_data
            
        def done(res):
            metrics, zones_data = res
            # Update values
            kpis["billed_val"].config(text=utils.format_currency(metrics["total_billed"]))
            kpis["collections_val"].config(text=utils.format_currency(metrics["total_collected"]))
            kpis["deficit_val"].config(text=utils.format_currency(metrics["total_outstanding"]))
            
            # Embed matplotlib bar chart
            if canvas_container[0]:
                canvas_container[0].get_tk_widget().destroy()
                
            fig = Figure(figsize=(5, 3.2), dpi=100)
            ax = fig.add_subplot(111)
            
            zones = [str(z["zone"]) for z in zones_data]
            billed = [z["billed_amount"] for z in zones_data]
            collected = [z["collected_amount"] for z in zones_data]
            
            import numpy as np
            x = np.arange(len(zones))
            width = 0.35
            
            ax.bar(x - width/2, billed, width, label="Billed (₹)", color="#1a3a6b")
            ax.bar(x + width/2, collected, width, label="Collected (₹)", color="#2a7a2a")
            
            ax.set_ylabel("Rupees (₹)")
            ax.set_title(f"Billing vs Collection per Zone (Cycle: {c_id})")
            ax.set_xticks(x)
            ax.set_xticklabels(zones)
            ax.legend()
            fig.tight_layout()
            
            canvas = FigureCanvasTkAgg(fig, master=chart_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill="both", expand=True)
            canvas_container[0] = canvas
            
        utils.run_in_thread(fetch, callback=done, widget=tab)

    ttk.Button(f_frame, text="📊 Run Analytics", command=generate_report).pack(side="left", padx=15)

# ----------------------------------------------------
# TAB 3: Zone Collections Breakdown Table
# ----------------------------------------------------
def build_zone_collection_tab(tab, fc, utils, be, admin):
    f_frame = ttk.Frame(tab, padding=5)
    f_frame.pack(fill="x", pady=(0, 10))
    
    ttk.Label(f_frame, text="Select Cycle:").pack(side="left", padx=5)
    cycle_var = tk.StringVar()
    cycle_cb = ttk.Combobox(f_frame, textvariable=cycle_var, width=18, state="readonly")
    cycle_cb.pack(side="left", padx=5)
    
    # Load cycles
    def refresh_cycles():
        def fetch():
            return fc.list_billing_cycles()
        def done(cycles):
            cycle_cb.config(values=[c["cycle_id"] for c in cycles])
            if cycles:
                cycle_cb.set(cycles[0]["cycle_id"])
        utils.run_in_thread(fetch, callback=done, widget=tab)
        
    refresh_cycles()
    
    tree_columns = ("zone", "total", "read", "cannot", "pending", "billed", "collected", "pct")
    tree = ttk.Treeview(tab, columns=tree_columns, show="headings")
    tree.heading("zone", text="Zone")
    tree.heading("total", text="Consumers")
    tree.heading("read", text="Read (Finalized)")
    tree.heading("cannot", text="Cannot Read")
    tree.heading("pending", text="Pending")
    tree.heading("billed", text="Billed Amount")
    tree.heading("collected", text="Collected Amount")
    tree.heading("pct", text="Collection %")
    
    tree.column("zone", width=50, anchor="center")
    tree.column("total", width=80, anchor="center")
    tree.column("read", width=100, anchor="center")
    tree.column("cannot", width=100, anchor="center")
    tree.column("pending", width=80, anchor="center")
    tree.column("billed", width=110, anchor="e")
    tree.column("collected", width=110, anchor="e")
    tree.column("pct", width=95, anchor="center")
    
    scrollbar = ttk.Scrollbar(tab, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    tree.pack(side="top", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    table_data = []
    
    def search():
        for r in tree.get_children():
            tree.delete(r)
            
        c_id = cycle_var.get()
        if not c_id:
            return
            
        def fetch():
            return fc.get_zone_collection_report(c_id)
            
        def done(zones):
            nonlocal table_data
            table_data = zones
            for z in zones:
                b = z["billed_amount"]
                c = z["collected_amount"]
                pct = (c / b * 100.0) if b > 0 else 0.0
                
                tree.insert("", "end", values=(
                    z["zone"],
                    z["total_consumers"],
                    z["read_consumers"],
                    z["cannot_read_consumers"],
                    z["pending_consumers"],
                    utils.format_currency(b),
                    utils.format_currency(c),
                    f"{pct:.1f} %"
                ))
        utils.run_in_thread(fetch, callback=done, widget=tab)

    ttk.Button(f_frame, text="🔍 Run breakdown", command=search).pack(side="left", padx=15)
    
    def export_excel():
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Save Zone Collections")
        if not path:
            return
        def run():
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Zone Collection"
            ws.append(["Zone", "Total Consumers", "Read (Finalized)", "Cannot Read", "Pending", "Billed Amount", "Collected Amount", "Collection %"])
            for z in table_data:
                b = z["billed_amount"]
                c = z["collected_amount"]
                pct = (c / b * 100) if b > 0 else 0
                ws.append([z["zone"], z["total_consumers"], z["read_consumers"], z["cannot_read_consumers"], z["pending_consumers"], b, c, f"{pct:.2f}%"])
            wb.save(path)
        def done(res): messagebox.showinfo("Success", "Spreadsheet exported.", parent=tab)
        utils.run_in_thread(run, callback=done, widget=tab)
        
    ttk.Button(f_frame, text="📤 Export Excel", command=export_excel).pack(side="right", padx=5)

# ----------------------------------------------------
# TAB 4: Outstanding Balance Report
# ----------------------------------------------------
def build_outstanding_tab(tab, fc, utils, be, admin):
    ttk.Label(tab, text="Defaulters List & Outstanding Balances", style="Header.TLabel", foreground="#1a3a6b").pack(anchor="w", pady=(0, 10))
    
    btn_frame = ttk.Frame(tab)
    btn_frame.pack(fill="x", pady=5)
    
    tree_columns = ("cin_no", "name", "zone", "category", "outstanding", "credit", "status")
    tree = ttk.Treeview(tab, columns=tree_columns, show="headings")
    tree.heading("cin_no", text="CIN")
    tree.heading("name", text="Consumer Name")
    tree.heading("zone", text="Zone")
    tree.heading("category", text="Category")
    tree.heading("outstanding", text="Outstanding")
    tree.heading("credit", text="Credit Balance")
    tree.heading("status", text="Status")
    
    tree.column("cin_no", width=110, anchor="center")
    tree.column("name", width=180)
    tree.column("zone", width=50, anchor="center")
    tree.column("category", width=95, anchor="center")
    tree.column("outstanding", width=110, anchor="e")
    tree.column("credit", width=100, anchor="e")
    tree.column("status", width=80, anchor="center")
    
    scrollbar = ttk.Scrollbar(tab, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    tree.pack(side="top", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    outstanding_cache = []
    
    def search():
        for r in tree.get_children():
            tree.delete(r)
        def fetch():
            return fc.get_outstanding_balance_report()
        def done(consumers):
            nonlocal outstanding_cache
            outstanding_cache = consumers
            for c in consumers:
                tree.insert("", "end", values=(
                    c["cin_no"],
                    c["name"],
                    c["zone"],
                    c["category"],
                    utils.format_currency(c["outstanding_balance"]),
                    utils.format_currency(c["credit_balance"]),
                    c["status"]
                ))
        utils.run_in_thread(fetch, callback=done, widget=tab)
        
    ttk.Button(btn_frame, text="🔄 Fetch Outstandings list", command=search).pack(side="left")
    
    def print_pdf_statement():
        sel = tree.selection()
        if not sel:
            messagebox.showerror("Error", "Please select a consumer row from the list.", parent=tab)
            return
        cin = sel[0] # use list item iid
        
        # Load profile details
        def run():
            c = fc.get_consumer(cin)
            if not c:
                raise FileNotFoundError(f"Consumer {cin} not found.")
            readings = fc.get_readings_for_cycle("", cin)
            payments = fc.get_payments_for_consumer(cin)
            adjustments = fc.get_adjustments_for_consumer(cin)
            replacements = fc.get_meter_replacement_history(cin)
            
            # Form rows
            r_rows = ""
            for r in readings[:10]: # limit to recent 10 for layout
                curr_val = f"{r.get('current_reading', 0.0):.2f}" if r.get('current_reading') is not None else ""
                cons_val = f"{r.get('consumption', 0.0):.2f}" if r.get('consumption') is not None else ""
                bill_breakdown = r.get("full_bill_breakdown")
                bill_tot = utils.format_currency(bill_breakdown.get("total_amount", 0.0)) if bill_breakdown else ""
                r_rows += f"<tr><td>{r.get('reading_date')}</td><td>{r.get('cycle_id')}</td><td>{r.get('reader_name')}</td><td>{r.get('previous_reading'):.2f}</td><td>{curr_val}</td><td>{cons_val}</td><td>{bill_tot}</td><td>{r.get('status')}</td></tr>"
                
            p_rows = ""
            for p in payments[:10]:
                p_rows += f"<tr><td>{p.get('receipt_number')}</td><td>{p.get('payment_date')}</td><td>{p.get('entry_date')}</td><td>{p.get('payment_mode')}</td><td>{p.get('emitra_key') or ''}</td><td>{p.get('received_by')}</td><td>{utils.format_currency(p.get('amount'))}</td></tr>"
                
            a_rows = ""
            for a in adjustments[:10]:
                a_rows += f"<tr><td>{utils.format_date(a.get('applied_at'))}</td><td>{a.get('type')}</td><td>{a.get('cycle_id') or ''}</td><td>{a.get('reason_note')}</td><td>{a.get('applied_by')}</td><td>{utils.format_currency(a.get('amount'))}</td></tr>"
                
            m_rows = ""
            for m in replacements[:10]:
                m_rows += f"<tr><td>{m.get('replacement_date')}</td><td>{m.get('old_meter_serial')}</td><td>{m.get('new_meter_serial')}</td><td>{m.get('new_initial_reading'):.2f}</td><td>{m.get('recorded_by')}</td></tr>"
                
            html_tmpl = utils.load_pdf_template("consumer_ledger")
            cin_no = c.get("cin_no", cin)
            html = html_tmpl.replace("{{cin_no}}", cin_no)\
                            .replace("{{meter_serial_no}}", c.get("meter_serial_no") or "")\
                            .replace("{{name}}", c["name"])\
                            .replace("{{category}}", c["category"])\
                            .replace("{{meter_size}}", c["meter_size"])\
                            .replace("{{apl_bpl}}", c.get("apl_bpl", "APL"))\
                            .replace("{{contact_number}}", str(c.get("contact_number") or ""))\
                            .replace("{{aadhaar_phed_no}}", c.get("aadhaar_phed_no") or "")\
                            .replace("{{address}}", c.get("address_area_location") or "")\
                            .replace("{{address_landmark}}", c.get("address_landmark") or "")\
                            .replace("{{address_pin_code}}", str(c.get("address_pin_code") or ""))\
                            .replace("{{address_latitude}}", str(c.get("address_latitude") or "0.00"))\
                            .replace("{{address_longitude}}", str(c.get("address_longitude") or "0.00"))\
                            .replace("{{consumer_status}}", c.get("consumer_status", "Active"))\
                            .replace("{{outstanding_balance}}", utils.format_currency(c.get("outstanding_balance", 0.0)))\
                            .replace("{{credit_balance}}", utils.format_currency(c.get("credit_balance", 0.0)))\
                            .replace("{{print_date}}", datetime.now().strftime("%d-%m-%Y"))\
                            .replace("{{print_time}}", datetime.now().strftime("%H:%M"))\
                            .replace("{{readings_rows}}", r_rows)\
                            .replace("{{payments_rows}}", p_rows)\
                            .replace("{{adjustments_rows}}", a_rows)\
                            .replace("{{replacements_rows}}", m_rows)
                            
            pdf_bytes = utils.render_pdf_to_bytes(html)
            temp_path = os.path.join(os.environ.get("TEMP", "."), f"ledger_{cin}.pdf")
            with open(temp_path, "wb") as f_out:
                f_out.write(pdf_bytes)
            utils.open_pdf(temp_path)
            
        utils.run_in_thread(run, widget=tab)

    ttk.Button(btn_frame, text="🖨️ Export PDF Ledger Statement", command=print_pdf_statement).pack(side="right")
    search()

# ----------------------------------------------------
# TAB 5: Skipped / Cannot Read list
# ----------------------------------------------------
def build_skipped_tab(tab, fc, utils, be, admin):
    f_frame = ttk.Frame(tab, padding=5)
    f_frame.pack(fill="x", pady=(0, 10))
    
    ttk.Label(f_frame, text="Select Cycle:").pack(side="left", padx=5)
    cycle_var = tk.StringVar()
    cycle_cb = ttk.Combobox(f_frame, textvariable=cycle_var, width=18, state="readonly")
    cycle_cb.pack(side="left", padx=5)
    
    # Load cycles
    def refresh_cycles():
        def fetch():
            return fc.list_billing_cycles()
        def done(cycles):
            cycle_cb.config(values=[c["cycle_id"] for c in cycles])
            if cycles:
                cycle_cb.set(cycles[0]["cycle_id"])
        utils.run_in_thread(fetch, callback=done, widget=tab)
        
    refresh_cycles()
    
    tree_columns = ("cin_no", "name", "zone", "reason", "reader", "date", "notes")
    tree = ttk.Treeview(tab, columns=tree_columns, show="headings")
    tree.heading("cin_no", text="CIN")
    tree.heading("name", text="Consumer Name")
    tree.heading("zone", text="Zone")
    tree.heading("reason", text="Reason")
    tree.heading("reader", text="Reader Name")
    tree.heading("date", text="Date")
    tree.heading("notes", text="Remarks Notes")
    
    tree.column("cin_no", width=110, anchor="center")
    tree.column("name", width=180)
    tree.column("zone", width=50, anchor="center")
    tree.column("reason", width=130)
    tree.column("reader", width=130)
    tree.column("date", width=100, anchor="center")
    tree.column("notes", width=200)
    
    scrollbar = ttk.Scrollbar(tab, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    tree.pack(side="top", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    table_data = []
    
    def search():
        for r in tree.get_children():
            tree.delete(r)
        c_id = cycle_var.get()
        if not c_id: return
        
        def fetch():
            return fc.get_skipped_readings_report(c_id)
        def done(data):
            nonlocal table_data
            table_data = data
            for r in data:
                tree.insert("", "end", values=(
                    r["cin_no"],
                    r["name"],
                    r["zone"],
                    r["reason"],
                    r["reader_name"],
                    r["reading_date"],
                    r.get("notes") or ""
                ))
        utils.run_in_thread(fetch, callback=done, widget=tab)
        
    ttk.Button(f_frame, text="🔍 Query Skipped List", command=search).pack(side="left", padx=15)

# ----------------------------------------------------
# TAB 6: Ledger statement views & bulk export
# ----------------------------------------------------
def build_ledger_statement_tab(tab, fc, utils, be, admin):
    ttk.Label(tab, text="Generate Consumer Ledger Statements", style="Header.TLabel", foreground="#1a3a6b").pack(anchor="w", pady=(0, 10))
    
    # Left search single, Right bulk zone export
    lf = ttk.LabelFrame(tab, text="Single Consumer Ledger Report", padding=15)
    lf.pack(fill="x", pady=10)
    
    # Search
    ttk.Label(lf, text="Search consumer (CIN or Meter Serial):").pack(side="left")
    search_ent = ttk.Entry(lf, width=20)
    search_ent.pack(side="left", padx=10)
    
    target_c = [None]
    c_status_lbl = ttk.Label(lf, text="", font=("Segoe UI", 9, "bold"))
    c_status_lbl.pack(side="left", padx=10)
    
    def lookup():
        q = search_ent.get().strip()
        if not q:
            return
        def fetch():
            res = fc.get_consumer(q)
            if not res:
                res = fc.get_consumer_by_meter_serial(q)
            return res
        def done(c):
            if c:
                target_c[0] = c
                c_status_lbl.config(text=f"Selected: {c['name']} (Zone {c['zone']})", foreground="#2a7a2a")
                exp_pdf_btn.config(state="normal")
            else:
                target_c[0] = None
                c_status_lbl.config(text="Not Found", foreground="red")
                exp_pdf_btn.config(state="disabled")
        utils.run_in_thread(fetch, callback=done, widget=tab)
        
    ttk.Button(lf, text="🔍 Match Profile", command=lookup).pack(side="left")
    
    def print_pdf_ledger():
        c = target_c[0]
        if not c:
            return
            
        def run():
            cin = c["cin_no"]
            readings = fc.get_readings_for_cycle("", cin)
            payments = fc.get_payments_for_consumer(cin)
            adjustments = fc.get_adjustments_for_consumer(cin)
            replacements = fc.get_meter_replacement_history(cin)
            
            r_rows = ""
            for r in readings:
                curr_val = f"{r.get('current_reading', 0.0):.2f}" if r.get('current_reading') is not None else ""
                cons_val = f"{r.get('consumption', 0.0):.2f}" if r.get('consumption') is not None else ""
                bill_breakdown = r.get("full_bill_breakdown")
                bill_tot = utils.format_currency(bill_breakdown.get("total_amount", 0.0)) if bill_breakdown else ""
                r_rows += f"<tr><td>{r.get('reading_date')}</td><td>{r.get('cycle_id')}</td><td>{r.get('reader_name')}</td><td>{r.get('previous_reading'):.2f}</td><td>{curr_val}</td><td>{cons_val}</td><td>{bill_tot}</td><td>{r.get('status')}</td></tr>"
                
            p_rows = ""
            for p in payments:
                p_rows += f"<tr><td>{p.get('receipt_number')}</td><td>{p.get('payment_date')}</td><td>{p.get('entry_date')}</td><td>{p.get('payment_mode')}</td><td>{p.get('emitra_key') or ''}</td><td>{p.get('received_by')}</td><td>{utils.format_currency(p.get('amount'))}</td></tr>"
                
            a_rows = ""
            for a in adjustments:
                a_rows += f"<tr><td>{utils.format_date(a.get('applied_at'))}</td><td>{a.get('type')}</td><td>{a.get('cycle_id') or ''}</td><td>{a.get('reason_note')}</td><td>{a.get('applied_by')}</td><td>{utils.format_currency(a.get('amount'))}</td></tr>"
                
            m_rows = ""
            for m in replacements:
                m_rows += f"<tr><td>{m.get('replacement_date')}</td><td>{m.get('old_meter_serial')}</td><td>{m.get('new_meter_serial')}</td><td>{m.get('new_initial_reading'):.2f}</td><td>{m.get('recorded_by')}</td></tr>"
                
            html_tmpl = utils.load_pdf_template("consumer_ledger")
            cin_no = c.get("cin_no", cin)
            html = html_tmpl.replace("{{cin_no}}", cin_no)\
                            .replace("{{meter_serial_no}}", c.get("meter_serial_no") or "")\
                            .replace("{{name}}", c["name"])\
                            .replace("{{category}}", c["category"])\
                            .replace("{{meter_size}}", c["meter_size"])\
                            .replace("{{apl_bpl}}", c.get("apl_bpl", "APL"))\
                            .replace("{{contact_number}}", str(c.get("contact_number") or ""))\
                            .replace("{{aadhaar_phed_no}}", c.get("aadhaar_phed_no") or "")\
                            .replace("{{address}}", c.get("address_area_location") or "")\
                            .replace("{{address_landmark}}", c.get("address_landmark") or "")\
                            .replace("{{address_pin_code}}", str(c.get("address_pin_code") or ""))\
                            .replace("{{address_latitude}}", str(c.get("address_latitude") or "0.00"))\
                            .replace("{{address_longitude}}", str(c.get("address_longitude") or "0.00"))\
                            .replace("{{consumer_status}}", c.get("consumer_status", "Active"))\
                            .replace("{{outstanding_balance}}", utils.format_currency(c.get("outstanding_balance", 0.0)))\
                            .replace("{{credit_balance}}", utils.format_currency(c.get("credit_balance", 0.0)))\
                            .replace("{{print_date}}", datetime.now().strftime("%d-%m-%Y"))\
                            .replace("{{print_time}}", datetime.now().strftime("%H:%M"))\
                            .replace("{{readings_rows}}", r_rows)\
                            .replace("{{payments_rows}}", p_rows)\
                            .replace("{{adjustments_rows}}", a_rows)\
                            .replace("{{replacements_rows}}", m_rows)
                            
            pdf_bytes = utils.render_pdf_to_bytes(html)
            temp_path = os.path.join(os.environ.get("TEMP", "."), f"ledger_view_{cin}.pdf")
            with open(temp_path, "wb") as f_out:
                f_out.write(pdf_bytes)
            utils.open_pdf(temp_path)
            
        utils.run_in_thread(run, widget=tab)

    exp_pdf_btn = ttk.Button(lf, text="🖨️ Open PDF Statement", command=print_pdf_ledger, state="disabled")
    exp_pdf_btn.pack(side="right")
    
    # Bulk Export Right
    rf = ttk.LabelFrame(tab, text="Bulk Zone Ledger Export (Excel Spreadsheet)", padding=15)
    rf.pack(fill="x", pady=10)
    
    ttk.Label(rf, text="Select Zone:").pack(side="left")
    zone_var = tk.StringVar(value="1")
    zone_cb = ttk.Combobox(rf, textvariable=zone_var, values=[str(z) for z in utils.ZONE_RANGE], width=6, state="readonly")
    zone_cb.pack(side="left", padx=10)
    
    progress = ttk.Progressbar(rf, orient="horizontal", mode="determinate", length=150)
    
    def bulk_excel_export():
        z = int(zone_var.get())
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Save Zone Ledgers Export")
        if not path:
            return
            
        progress.pack(side="left", padx=15)
        progress["value"] = 10
        
        def run():
            consumers = fc.list_consumers({"zone": z})
            wb = openpyxl.Workbook()
            # remove default sheet
            wb.remove(wb.active)
            
            tot = len(consumers)
            for idx, c in enumerate(consumers):
                cin = c["cin_no"]
                ws = wb.create_sheet(title=cin)
                
                # Write layout keys
                ws.append(["Rajasthan Jal Board — Account Ledger Summary"])
                ws.append(["CIN Number", cin, "Meter Serial", c.get("meter_serial_no")])
                ws.append(["Consumer Name", c["name"], "Zone", c["zone"]])
                ws.append(["Billing Category", c["category"], "Meter Size", c["meter_size"]])
                ws.append(["Outstanding Balance", c.get("outstanding_balance", 0.0), "Credit Balance", c.get("credit_balance", 0.0)])
                ws.append([])
                
                # Write readings
                ws.append(["Readings History Logs"])
                ws.append(["Date", "Cycle ID", "Reader", "Previous Reading", "Current Reading", "Consumption", "Billed Total", "Status"])
                readings = fc.get_readings_for_cycle("", cin)
                for r in readings:
                    bill_breakdown = r.get("full_bill_breakdown")
                    bill_tot = bill_breakdown.get("total_amount", 0.0) if bill_breakdown else 0.0
                    ws.append([r.get("reading_date"), r.get("cycle_id"), r.get("reader_name"), r.get("previous_reading"), r.get("current_reading"), r.get("consumption"), bill_tot, r.get("status")])
                ws.append([])
                
                # Write payments
                ws.append(["Payments In Logs"])
                ws.append(["Receipt Number", "Payment Date", "Payment Mode", "E-Mitra Key", "Collected By", "Amount Paid"])
                payments = fc.get_payments_for_consumer(cin)
                for p in payments:
                    ws.append([p.get("receipt_number"), p.get("payment_date"), p.get("payment_mode"), p.get("emitra_key") or "", p.get("received_by"), p.get("amount")])
                    
                # Update progress bar
                progress["value"] = 10 + int((idx + 1) / tot * 80)
                
            wb.save(path)
            progress["value"] = 100
            
        def done(res):
            progress.pack_forget()
            messagebox.showinfo("Success", f"Zone ledger spreadsheet created at:\n{path}", parent=tab)
            
        utils.run_in_thread(run, callback=done, widget=tab)

    ttk.Button(rf, text="📤 Export All Ledgers to Excel", command=bulk_excel_export).pack(side="right")

# ----------------------------------------------------
# TAB 7: Database Full Backup tools
# ----------------------------------------------------
def build_backup_tab(tab, fc, utils, be, admin):
    ttk.Label(tab, text="Database Security & System Backups", style="Header.TLabel", foreground="#1a3a6b").pack(anchor="w", pady=(0, 10))
    
    desc = "⚠️ ATTENTION: Exporting full backup downloads all records from Firebase (Consumers, Readings, Payments, Logs) " \
           "and writes them into separate sheets inside a single Excel Workbook. " \
           "This represents a resource intensive request. Run only when necessary."
    ttk.Label(tab, text=desc, foreground="#c0392b", font=("Segoe UI", 9, "bold"), wrap=500).pack(anchor="w", pady=15)
    
    prog_bar = ttk.Progressbar(tab, orient="horizontal", mode="determinate", length=300)
    
    def run_backup():
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Save Database Backup")
        if not path:
            return
            
        prog_bar.pack(anchor="w", pady=10)
        prog_bar["value"] = 10
        
        def run():
            backup_data = fc.export_full_data_backup()
            prog_bar["value"] = 50
            
            wb = openpyxl.Workbook()
            # remove default
            wb.remove(wb.active)
            
            for c_name, rows in backup_data.items():
                ws = wb.create_sheet(title=c_name[:30]) # excel tab limit is 31 chars
                if not rows:
                    ws.append(["No records found in this collection."])
                    continue
                # Get headers dynamically
                headers = list(rows[0].keys())
                ws.append(headers)
                
                # Headers style
                header_font = openpyxl.styles.Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
                header_fill = openpyxl.styles.PatternFill(start_color="1A3A6B", end_color="1A3A6B", fill_type="solid")
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    
                for r in rows:
                    row_vals = [r.get(h) for h in headers]
                    ws.append(row_vals)
                    
                for col in ws.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
                    
            wb.save(path)
            prog_bar["value"] = 100
            
        def done(res):
            prog_bar.pack_forget()
            messagebox.showinfo("Success", f"Database system backup saved successfully at:\n{path}", parent=tab)
            
        def fail(err):
            prog_bar.pack_forget()
            messagebox.showerror("Error", f"Failed compiling database backup:\n{err}", parent=tab)
            
        utils.run_in_thread(run, callback=done, error_callback=fail, widget=tab)

    ttk.Button(tab, text="💾 Run Full System Backup", command=run_backup).pack(anchor="w", pady=10)
