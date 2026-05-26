import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
from datetime import datetime

def get_frame(parent, fc, utils, be, admin) -> ttk.Frame:
    frame = ttk.Frame(parent)
    
    # Title and Refresh row
    header = ttk.Frame(frame)
    header.pack(fill="x", pady=(0, 20))
    ttk.Label(header, text="Meter Readings & Queries", style="Title.TLabel").pack(side="left")
    ttk.Button(header, text="🔄 Refresh Readings", style="Primary.TButton", command=lambda: view_tab.perform_search(refresh=True)).pack(side="right")

    notebook = ttk.Notebook(frame)
    notebook.pack(fill="both", expand=True)
    
    # Tab 1: View Readings
    view_tab = ttk.Frame(notebook, padding=10)
    notebook.add(view_tab, text="📖 View Readings")
    
    # Tab 2: Pending Correction Queries
    pending_tab = ttk.Frame(notebook, padding=10)
    notebook.add(pending_tab, text="⚠️ Pending Corrections")
    
    # Tab 3: All Correction Queries
    all_queries_tab = ttk.Frame(notebook, padding=10)
    notebook.add(all_queries_tab, text="📜 All Queries Log")
    
    build_view_tab(view_tab, fc, utils, be, admin)
    build_pending_tab(pending_tab, fc, utils, be, admin, notebook, 1)
    build_all_queries_tab(all_queries_tab, fc, utils, be, admin)
    
    return frame

# ----------------------------------------------------
# TAB 1: View Readings & Overrides
# ----------------------------------------------------
def build_view_tab(tab, fc, utils, be, admin):
    # Filters
    f_frame = ttk.Frame(tab, padding=5)
    f_frame.pack(fill="x", pady=(0, 10))
    
    ttk.Label(f_frame, text="Cycle:").pack(side="left", padx=5)
    cycle_var = tk.StringVar(value="All")
    cycle_cb = ttk.Combobox(f_frame, textvariable=cycle_var, width=15, state="readonly")
    cycle_cb.pack(side="left", padx=5)
    
    ttk.Label(f_frame, text="Zone:").pack(side="left", padx=5)
    zone_var = tk.StringVar(value="All")
    zone_cb = ttk.Combobox(f_frame, textvariable=zone_var, values=["All"] + [str(z) for z in utils.ZONE_RANGE], width=6, state="readonly")
    zone_cb.pack(side="left", padx=5)
    
    ttk.Label(f_frame, text="Status:").pack(side="left", padx=5)
    status_var = tk.StringVar(value="All")
    status_cb = ttk.Combobox(f_frame, textvariable=status_var, values=["All", "finalized", "skipped", "Anomaly Flagged"], width=12, state="readonly")
    status_cb.pack(side="left", padx=5)
    
    tree_columns = ("date", "cin_no", "reader", "prev", "curr", "consumption", "bill", "status", "edited")
    tree = ttk.Treeview(tab, columns=tree_columns, show="headings")
    tree.heading("date", text="Date & Time")
    tree.heading("cin_no", text="CIN No")
    tree.heading("reader", text="Reader")
    tree.heading("prev", text="Prev (KL)")
    tree.heading("curr", text="Curr (KL)")
    tree.heading("consumption", text="Usage (KL)")
    tree.heading("bill", text="Bill Total")
    tree.heading("status", text="Status")
    tree.heading("edited", text="Override")
    
    tree.column("date", width=110, anchor="center")
    tree.column("cin_no", width=100, anchor="center")
    tree.column("reader", width=120)
    tree.column("prev", width=70, anchor="e")
    tree.column("curr", width=70, anchor="e")
    tree.column("consumption", width=70, anchor="e")
    tree.column("bill", width=90, anchor="e")
    tree.column("status", width=80, anchor="center")
    tree.column("edited", width=60, anchor="center")
    
    scrollbar = ttk.Scrollbar(tab, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    tree.pack(side="top", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    readings_list = []
    
    def refresh_cycles_filter():
        def fetch():
            return fc.list_billing_cycles()
        def done(cycles):
            cycle_cb.config(values=["All"] + [c["cycle_id"] for c in cycles])
        utils.run_in_thread(fetch, callback=done, widget=tab)
        
    refresh_cycles_filter()
    
    def perform_search(refresh=False):
        for r in tree.get_children():
            tree.delete(r)
            
        c_filter = cycle_var.get()
        z_filter = zone_var.get()
        s_filter = status_var.get()
        
        def fetch():
            # Firestore query limitations: we fetch readings and evaluate local filters
            # if cycle is selected, filter by cycle to minimize reads
            cycle_id = c_filter if c_filter != "All" else ""
            readings = fc.get_readings_for_cycle(cycle_id, use_cache=not refresh)
            return readings
            
        def done(readings):
            nonlocal readings_list
            readings_list = readings
            
            # Local client side filters
            filtered = []
            for r in readings:
                # Resolve consumer info for Zone
                cons = fc.get_consumer(r["cin_no"])
                if not cons:
                    continue
                    
                if z_filter != "All" and int(cons.get("zone", 0)) != int(z_filter):
                    continue
                    
                if s_filter == "skipped" and r.get("status") != "skipped":
                    continue
                if s_filter == "finalized" and r.get("status") != "finalized":
                    continue
                if s_filter == "Anomaly Flagged" and not r.get("anomaly_flagged", False):
                    continue
                    
                filtered.append((r, cons))
                
            for r, c in filtered:
                sub_at = r.get("submitted_at")
                dt_str = utils.format_date(sub_at)
                
                curr_val = f"{r.get('current_reading', 0.0):.2f}" if r.get("current_reading") is not None else ""
                cons_val = f"{r.get('consumption', 0.0):.2f}" if r.get("consumption") is not None else ""
                
                bill_breakdown = r.get("full_bill_breakdown")
                bill_tot = utils.format_currency(bill_breakdown.get("total_amount", 0.0)) if bill_breakdown else ""
                
                tree.insert("", "end", values=(
                    dt_str,
                    r["cin_no"],
                    r.get("reader_name", "Field Reader"),
                    f"{r.get('previous_reading', 0.0):.2f}",
                    curr_val,
                    cons_val,
                    bill_tot,
                    r.get("status", ""),
                    "Yes" if r.get("edited_by_admin", False) else "No"
                ), iid=r["reading_id"])
                
        utils.run_in_thread(fetch, callback=done, widget=tab)

    ttk.Button(f_frame, text="🔍 Filter Readings", command=perform_search).pack(side="left", padx=15)
    
    # Bulk Export / Import buttons
    def download_readings():
        c_filter = cycle_var.get()
        if c_filter == "All":
            messagebox.showerror("Error", "Please select a specific Billing Cycle to export for corrections.", parent=tab)
            return
            
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")], title="Save Readings Export")
        if not path:
            return
            
        def run():
            readings = fc.get_readings_for_cycle(c_filter)
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Corrections Log"
            
            headers = ["reading_id", "cin_no", "cycle_id", "reader_name", "previous_reading", "current_reading", "notes"]
            ws.append(headers)
            
            for r in readings:
                ws.append([
                    r["reading_id"],
                    r["cin_no"],
                    r["cycle_id"],
                    r.get("reader_name", ""),
                    r.get("previous_reading", 0.0),
                    r.get("current_reading"),
                    r.get("notes", "")
                ])
            wb.save(path)
            
        def done(res):
            messagebox.showinfo("Success", f"Readings saved successfully to:\n{path}", parent=tab)
            
        utils.run_in_thread(run, callback=done, widget=tab)
        
    ttk.Button(f_frame, text="📤 Export Cycle to Excel", command=download_readings).pack(side="right", padx=5)
    
    def import_corrections():
        path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")], title="Select Readings Correction File")
        if not path:
            return
            
        def run():
            wb = openpyxl.load_workbook(path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            
            updates = []
            headers = [str(c).strip() for c in rows[0]]
            
            for r_idx, row in enumerate(rows[1:]):
                if not row or not any(row):
                    continue
                payload = {}
                for idx, h in enumerate(headers):
                    payload[h] = row[idx]
                    
                # We need reading_id and current_reading
                if not payload.get("reading_id") or payload.get("current_reading") is None:
                    continue
                    
                updates.append({
                    "reading_id": payload["reading_id"],
                    "current_reading": float(payload["current_reading"]),
                    "notes": payload.get("notes", "Bulk update")
                })
                
            res = fc.bulk_update_readings(updates, admin["name"])
            return res
            
        def done(res):
            success = res["success"]
            errors = res["errors"]
            msg = f"Bulk corrections loaded successfully!\n\nUpdated rows: {success}"
            if errors:
                msg += f"\nFailed edits: {len(errors)}"
                messagebox.showwarning("Warning", msg, parent=tab)
            else:
                messagebox.showinfo("Success", msg, parent=tab)
            perform_search()
            
        def fail(err):
            messagebox.showerror("Error", f"Bulk corrections import failed:\n{err}", parent=tab)
            
        utils.run_in_thread(run, callback=done, error_callback=fail, widget=tab)

    ttk.Button(f_frame, text="📥 Import Corrected Excel", command=import_corrections).pack(side="right", padx=5)

    # Double click details popup
    def on_row_double_click(event):
        sel = tree.selection()
        if not sel:
            return
        reading_id = sel[0]
        show_reading_detail_popup(reading_id, fc, utils, be, admin, tab, perform_search)
        
    tree.bind("<Double-1>", on_row_double_click)
    tab.perform_search = perform_search

def show_reading_detail_popup(reading_id, fc, utils, be, admin, parent, refresh_callback):
    popup = tk.Toplevel(parent)
    popup.title(f"Reading Breakdown — {reading_id}")
    popup.geometry("600x480")
    popup.grab_set()
    
    main_f = ttk.Frame(popup, padding=15)
    main_f.pack(fill="both", expand=True)
    
    def load_details():
        for w in main_f.winfo_children():
            w.destroy()
            
        ttk.Label(main_f, text="Reading Details", font=("Segoe UI", 12, "bold"), foreground="#1a3a6b").pack(anchor="w")
        
        def fetch():
            r = fc.get_reading(reading_id)
            c = fc.get_consumer(r["cin_no"]) if r else None
            return r, c
            
        def done(res):
            r, c = res
            if not r:
                messagebox.showerror("Error", "Reading not found.", parent=popup)
                popup.destroy()
                return
                
            # Details grid
            grid = ttk.Frame(main_f)
            grid.pack(fill="x", pady=10)
            
            # Left column
            ttk.Label(grid, text="CIN Number:", font=("Segoe UI", 9, "bold")).grid(row=0, column=0, sticky="w", pady=3)
            ttk.Label(grid, text=r["cin_no"]).grid(row=0, column=1, sticky="w", pady=3, padx=10)
            
            ttk.Label(grid, text="Consumer Name:", font=("Segoe UI", 9, "bold")).grid(row=1, column=0, sticky="w", pady=3)
            ttk.Label(grid, text=c["name"] if c else "Unknown").grid(row=1, column=1, sticky="w", pady=3, padx=10)
            
            ttk.Label(grid, text="Meter Serial No:", font=("Segoe UI", 9, "bold")).grid(row=2, column=0, sticky="w", pady=3)
            ttk.Label(grid, text=c.get("meter_serial_no", "") if c else "").grid(row=2, column=1, sticky="w", pady=3, padx=10)
            
            ttk.Label(grid, text="Reading Status:", font=("Segoe UI", 9, "bold")).grid(row=3, column=0, sticky="w", pady=3)
            ttk.Label(grid, text=r.get("status", "finalized")).grid(row=3, column=1, sticky="w", pady=3, padx=10)
            
            # Right column
            ttk.Label(grid, text="Reader Name:", font=("Segoe UI", 9, "bold")).grid(row=0, column=2, sticky="w", pady=3, padx=20)
            ttk.Label(grid, text=r.get("reader_name", "")).grid(row=0, column=3, sticky="w", pady=3)
            
            ttk.Label(grid, text="Previous Reading:", font=("Segoe UI", 9, "bold")).grid(row=1, column=2, sticky="w", pady=3, padx=20)
            ttk.Label(grid, text=f"{r.get('previous_reading', 0.0):.2f} KL").grid(row=1, column=3, sticky="w", pady=3)
            
            ttk.Label(grid, text="Current Reading:", font=("Segoe UI", 9, "bold")).grid(row=2, column=2, sticky="w", pady=3, padx=20)
            curr = r.get("current_reading")
            curr_str = f"{curr:.2f} KL" if curr is not None else "— (Skipped)"
            ttk.Label(grid, text=curr_str).grid(row=2, column=3, sticky="w", pady=3)
            
            ttk.Label(grid, text="Consumption:", font=("Segoe UI", 9, "bold")).grid(row=3, column=2, sticky="w", pady=3, padx=20)
            cons = r.get("consumption")
            cons_str = f"{cons:.2f} KL" if cons is not None else "— (Skipped)"
            ttk.Label(grid, text=cons_str).grid(row=3, column=3, sticky="w", pady=3)
            
            # Bill Breakdown Box
            bb = r.get("full_bill_breakdown")
            if bb:
                ttk.Label(main_f, text="Bill Breakdown Map", font=("Segoe UI", 10, "bold"), foreground="#1a3a6b").pack(anchor="w", pady=(10, 5))
                bb_txt = tk.Text(main_f, font=("Courier New", 9), height=7, bg="#fbfbfb")
                bb_txt.pack(fill="x", pady=5)
                
                # Format map keys cleanly
                txt = f"Water Charges Amount   : {utils.format_currency(bb.get('water_charge'))}\n"
                if bb.get("minimum_charge_applied"):
                    txt += f"  *Applied Minimum     : {utils.format_currency(bb.get('minimum_charge_amount'))}\n"
                txt += f"Fixed Renovation Charge: {utils.format_currency(bb.get('fixed_charge'))}\n"
                txt += f"Meter Service Charge   : {utils.format_currency(bb.get('meter_service_charge'))}\n"
                txt += f"IDS Surcharge Amount   : {utils.format_currency(bb.get('ids_charge'))} ({bb.get('ids_rate_pct')}%)\n"
                txt += f"Previous Outstanding   : {utils.format_currency(bb.get('previous_outstanding'))}\n"
                txt += f"Credit Applied Amount  : {utils.format_currency(bb.get('credit_applied'))}\n"
                txt += f"Late Payment Surcharge : {utils.format_currency(bb.get('lps_amount'))} (Type: {bb.get('lps_type')})\n"
                txt += f"NET RUPEES CEILING DUE : {utils.format_currency(bb.get('total_amount'))}\n"
                
                bb_txt.insert("end", txt)
                bb_txt.config(state="disabled")
                
            # Override button
            if r.get("status") == "finalized":
                def open_override_dlg():
                    show_admin_override_dialog(r, c, fc, utils, be, admin, popup, lambda: [load_details(), refresh_callback()])
                ttk.Button(main_f, text="🛠️ Admin Override Edit", command=open_override_dlg).pack(anchor="e", pady=10)
                
        utils.run_in_thread(fetch, callback=done, widget=popup)
        
    load_details()

def show_admin_override_dialog(reading, consumer, fc, utils, be, admin, parent, success_callback):
    dlg = tk.Toplevel(parent)
    dlg.title("Admin Reading Override")
    dlg.geometry("520x420")
    dlg.grab_set()
    
    f = ttk.Frame(dlg, padding=15)
    f.pack(fill="both", expand=True)
    
    ttk.Label(f, text="Admin Reading Override Edit", style="Header.TLabel", foreground="#1a3a6b").pack(anchor="w")
    ttk.Label(f, text=f"Modifying record {reading['reading_id']} for consumer {reading['cin_no']}").pack(anchor="w", pady=(0, 15))
    
    grid = ttk.Frame(f)
    grid.pack(fill="x")
    
    ttk.Label(grid, text="Previous Reading (KL):").grid(row=0, column=0, sticky="w", pady=5)
    ttk.Label(grid, text=f"{reading['previous_reading']:.2f} KL", font=("Segoe UI", 9, "bold")).grid(row=0, column=1, sticky="w", pady=5, padx=10)
    
    ttk.Label(grid, text="Submitted Reading (KL):").grid(row=1, column=0, sticky="w", pady=5)
    ttk.Label(grid, text=f"{reading['current_reading']:.2f} KL").grid(row=1, column=1, sticky="w", pady=5, padx=10)
    
    ttk.Label(grid, text="New Current Reading (KL):*").grid(row=2, column=0, sticky="w", pady=5)
    new_curr_ent = ttk.Entry(grid, width=15)
    new_curr_ent.grid(row=2, column=1, sticky="w", pady=5, padx=10)
    new_curr_ent.insert(0, str(reading["current_reading"]))
    
    ttk.Label(grid, text="Required Reason Note:*").grid(row=3, column=0, sticky="nw", pady=5)
    reason_txt = tk.Text(grid, height=3, width=30, font=("Segoe UI", 9))
    reason_txt.grid(row=3, column=1, sticky="w", pady=5, padx=10)
    
    # Bill comparison display panel
    comp_lbl = ttk.Label(f, text="", font=("Courier New", 9), background="#fafafa", relief="dashed", borderwidth=1, padding=5)
    comp_lbl.pack(fill="x", pady=10)
    
    def simulate_comparison(*args):
        val_str = new_curr_ent.get().strip()
        try:
            val = float(val_str)
            prev = float(reading["previous_reading"])
            new_cons = val - prev
            if new_cons < 0:
                comp_lbl.config(text="Error: New reading is less than previous reading.", foreground="red")
                return
                
            rates = fc.get_charges_config()
            cycle = fc.get_billing_cycle(reading["cycle_id"])
            last_pay_d = utils.parse_date(cycle["last_payment_date"]) if cycle else None
            
            # Old bill total
            old_bill = reading["full_bill_breakdown"]["total_amount"]
            
            # Simulate new bill
            sim_res = be.calculate_bill(
                consumption_kl=new_cons,
                consumer=consumer,
                rates=rates,
                previous_outstanding=float(consumer["outstanding_balance"]) - old_bill,
                credit_balance=float(consumer["credit_balance"]),
                last_payment_date=last_pay_d,
                payment_date=date.today()
            )
            
            txt = f"Bill Comparison Preview:\n" \
                  f"• Old Consumption: {reading['consumption']:.2f} KL ➔ New: {new_cons:.2f} KL\n" \
                  f"• Old Bill Total : {utils.format_currency(old_bill)} ➔ New: {utils.format_currency(sim_res['total_amount'])}"
            comp_lbl.config(text=txt, foreground="black")
        except Exception as e:
            comp_lbl.config(text="Simulation failed.", foreground="red")
            
    new_curr_ent.bind("<KeyRelease>", simulate_comparison)
    simulate_comparison()
    
    def save_override():
        new_val_str = new_curr_ent.get().strip()
        reason = reason_txt.get("1.0", "end").strip()
        
        if not new_val_str or not reason:
            messagebox.showerror("Error", "All fields are required.", parent=dlg)
            return
            
        try:
            new_val = float(new_val_str)
        except ValueError:
            messagebox.showerror("Error", "Invalid current reading value.", parent=dlg)
            return
            
        confirm = messagebox.askyesno("Confirm Override", "This will rewrite the reading and update the consumer outstanding balance. Confirm?", parent=dlg)
        if not confirm:
            return
            
        payload = {
            "current_reading": new_val,
            "notes": f"Admin Override: {reason}"
        }
        
        def run():
            fc.admin_update_reading(reading["reading_id"], payload, admin["name"])
            
        def done(res):
            messagebox.showinfo("Success", "Reading overridden successfully.", parent=parent)
            dlg.destroy()
            success_callback()
            
        def fail(err):
            messagebox.showerror("Override Failed", f"Error:\n{err}", parent=dlg)
            
        utils.run_in_thread(run, callback=done, error_callback=fail, widget=dlg)
        
    ttk.Button(f, text="🛠️ Apply Override", command=save_override).pack(anchor="e", pady=10)

# ----------------------------------------------------
# TAB 2: Pending Correction Queries
# ----------------------------------------------------
def build_pending_tab(tab, fc, utils, be, admin, notebook, tab_idx):
    for w in tab.winfo_children():
        w.destroy()
        
    ttk.Label(tab, text="Pending Correction Queries", style="Header.TLabel", foreground="#c0392b").pack(anchor="w", pady=(0, 10))
    
    tree_columns = ("date", "cin_no", "name", "reader", "submitted", "requested", "reason")
    tree = ttk.Treeview(tab, columns=tree_columns, show="headings", height=8)
    tree.heading("date", text="Date")
    tree.heading("cin_no", text="CIN")
    tree.heading("name", text="Consumer Name")
    tree.heading("reader", text="Reader Name")
    tree.heading("submitted", text="Submitted (KL)")
    tree.heading("requested", text="Requested (KL)")
    tree.heading("reason", text="Reason / Note")
    
    tree.column("date", width=100, anchor="center")
    tree.column("cin_no", width=100, anchor="center")
    tree.column("name", width=130)
    tree.column("reader", width=120)
    tree.column("submitted", width=90, anchor="e")
    tree.column("requested", width=90, anchor="e")
    tree.column("reason", width=220)
    
    scrollbar = ttk.Scrollbar(tab, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    tree.pack(side="top", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    queries_cache = {}
    
    def refresh_pending():
        for r in tree.get_children():
            tree.delete(r)
            
        def fetch():
            return fc.get_pending_correction_queries()
            
        def done(queries):
            count = len(queries)
            # Update notebook tab badge count
            text_str = f"⚠️ Pending Corrections ({count})" if count > 0 else "Pending Corrections"
            notebook.tab(tab_idx, text=text_str)
            
            for q in queries:
                queries_cache[q["query_id"]] = q
                c_snap = q.get("consumer_info_snapshot", {})
                tree.insert("", "end", values=(
                    utils.format_date(q.get("created_at")),
                    q["cin_no"],
                    c_snap.get("name", "Unknown"),
                    q.get("reader_name", "Field Reader"),
                    f"{q.get('previous_reading', 0.0):.2f}",
                    f"{q.get('requested_corrected_reading', 0.0):.2f}",
                    q.get("reason", "")
                ), iid=q["query_id"])
                
        utils.run_in_thread(fetch, callback=done, widget=tab)
        
    def on_row_double_click(event):
        sel = tree.selection()
        if not sel:
            return
        query_id = sel[0]
        q = queries_cache.get(query_id)
        if not q:
            return
            
        # Details & action popup
        dlg = tk.Toplevel(tab)
        dlg.title(f"Review Query — {query_id}")
        dlg.geometry("450x380")
        dlg.grab_set()
        
        f = ttk.Frame(dlg, padding=15)
        f.pack(fill="both", expand=True)
        
        ttk.Label(f, text="Review Correction Request", font=("Segoe UI", 12, "bold"), foreground="#1a3a6b").pack(anchor="w", pady=(0, 10))
        
        # Details text
        details = f"• Consumer CIN: {q['cin_no']}\n" \
                  f"• Reader Name : {q.get('reader_name')} (Emp ID: {q.get('reader_employee_id')})\n" \
                  f"• Prev Reading: {q.get('previous_reading'):.2f} KL\n" \
                  f"• Submitted   : {q.get('submitted_reading'):.2f} KL\n" \
                  f"• Requested   : {q.get('requested_corrected_reading'):.2f} KL\n" \
                  f"• Reason Given: {q.get('reason')}"
                  
        ttk.Label(f, text=details, font=("Segoe UI", 10), justify="left").pack(anchor="w", pady=10)
        
        ttk.Label(f, text="Rejection note (Required only if rejecting):", font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=2)
        reject_txt = tk.Text(f, height=3, width=45, font=("Segoe UI", 9))
        reject_txt.pack(pady=5)
        
        # Approve button
        def approve():
            confirm = messagebox.askyesno("Confirm Approve", "This will rewrite the verified reading and update outstanding balances. Confirm?", parent=dlg)
            if not confirm:
                return
                
            def run():
                fc.approve_correction_query(query_id, admin["name"])
                
            def done(res):
                messagebox.showinfo("Approved", "Reading correction has been applied successfully.", parent=tab)
                dlg.destroy()
                refresh_pending()
                
            def fail(err):
                messagebox.showerror("Error", f"Action failed:\n{err}", parent=dlg)
                
            utils.run_in_thread(run, callback=done, error_callback=fail, widget=dlg)
            
        # Reject button
        def reject():
            note = reject_txt.get("1.0", "end").strip()
            if not note:
                messagebox.showerror("Error", "Please enter a justification note for rejecting the correction query.", parent=dlg)
                return
                
            def run():
                fc.reject_correction_query(query_id, note, admin["name"])
                
            def done(res):
                messagebox.showinfo("Rejected", "Correction request rejected successfully.", parent=tab)
                dlg.destroy()
                refresh_pending()
                
            def fail(err):
                messagebox.showerror("Error", f"Action failed:\n{err}", parent=dlg)
                
            utils.run_in_thread(run, callback=done, error_callback=fail, widget=dlg)

        btn_f = ttk.Frame(f)
        btn_f.pack(fill="x", pady=15)
        ttk.Button(btn_f, text="✅ Approve Correction", command=approve).pack(side="left", padx=5)
        ttk.Button(btn_f, text="❌ Reject Request", command=reject).pack(side="right", padx=5)
        
    tree.bind("<Double-1>", on_row_double_click)
    refresh_pending()

# ----------------------------------------------------
# TAB 3: All Correction Queries Log
# ----------------------------------------------------
def build_all_queries_tab(tab, fc, utils, be, admin):
    f_frame = ttk.Frame(tab, padding=5)
    f_frame.pack(fill="x", pady=(0, 10))
    
    ttk.Label(f_frame, text="Filter by Cycle ID:").pack(side="left", padx=5)
    cycle_var = tk.StringVar(value="All")
    cycle_cb = ttk.Combobox(f_frame, textvariable=cycle_var, width=15, state="readonly")
    cycle_cb.pack(side="left", padx=5)
    
    def refresh_cycles():
        def fetch():
            return fc.list_billing_cycles()
        def done(cycles):
            cycle_cb.config(values=["All"] + [c["cycle_id"] for c in cycles])
        utils.run_in_thread(fetch, callback=done, widget=tab)
        
    refresh_cycles()
    
    tree_columns = ("date", "cin_no", "reader", "submitted", "requested", "status", "resolved")
    tree = ttk.Treeview(tab, columns=tree_columns, show="headings")
    tree.heading("date", text="Created At")
    tree.heading("cin_no", text="CIN")
    tree.heading("reader", text="Reader Name")
    tree.heading("submitted", text="Submitted (KL)")
    tree.heading("requested", text="Requested (KL)")
    tree.heading("status", text="Status")
    tree.heading("resolved", text="Resolved At")
    
    tree.column("date", width=110, anchor="center")
    tree.column("cin_no", width=100, anchor="center")
    tree.column("reader", width=130)
    tree.column("submitted", width=95, anchor="e")
    tree.column("requested", width=95, anchor="e")
    tree.column("status", width=80, anchor="center")
    tree.column("resolved", width=110, anchor="center")
    
    scrollbar = ttk.Scrollbar(tab, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=scrollbar.set)
    tree.pack(side="top", fill="both", expand=True)
    scrollbar.pack(side="right", fill="y")
    
    queries_cache = {}
    
    def perform_search():
        for r in tree.get_children():
            tree.delete(r)
            
        c_filter = cycle_var.get()
        
        def fetch():
            cycle_id = c_filter if c_filter != "All" else None
            return fc.get_all_correction_queries(cycle_id)
            
        def done(queries):
            for q in queries:
                queries_cache[q["query_id"]] = q
                tree.insert("", "end", values=(
                    utils.format_date(q.get("created_at")),
                    q["cin_no"],
                    q.get("reader_name"),
                    f"{q.get('submitted_reading', 0.0):.2f}",
                    f"{q.get('requested_corrected_reading', 0.0):.2f}",
                    q.get("status"),
                    utils.format_date(q.get("resolved_at"))
                ), iid=q["query_id"])
                
        utils.run_in_thread(fetch, callback=done, widget=tab)

    ttk.Button(f_frame, text="🔍 Filter Queries", command=perform_search).pack(side="left", padx=15)
    
    def on_row_double_click(event):
        sel = tree.selection()
        if not sel:
            return
        query_id = sel[0]
        q = queries_cache.get(query_id)
        
        dlg = tk.Toplevel(tab)
        dlg.title(f"Correction Log Info — {query_id}")
        dlg.geometry("420x300")
        dlg.grab_set()
        
        f = ttk.Frame(dlg, padding=15)
        f.pack(fill="both", expand=True)
        
        ttk.Label(f, text="Correction Query Details", font=("Segoe UI", 12, "bold"), foreground="#1a3a6b").pack(anchor="w", pady=(0, 10))
        
        details = f"• CIN No: {q['cin_no']}\n" \
                  f"• Submitted Reading: {q.get('submitted_reading'):.2f} KL\n" \
                  f"• Requested Corrected: {q.get('requested_corrected_reading'):.2f} KL\n" \
                  f"• Status: {q.get('status').upper()}\n" \
                  f"• Reason Given: {q.get('reason')}\n"
                  
        if q.get("status") == "rejected":
            details += f"• Rejection Note: {q.get('rejection_note', 'None')}\n"
            
        details += f"• Resolved Date: {utils.format_date(q.get('resolved_at'))}"
        
        ttk.Label(f, text=details, font=("Segoe UI", 10), justify="left").pack(anchor="w", pady=5)
        
    tree.bind("<Double-1>", on_row_double_click)
    perform_search()
